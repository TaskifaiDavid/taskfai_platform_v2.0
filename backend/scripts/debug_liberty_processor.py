"""
Debug Liberty Processor

Tests Liberty processor against actual file to identify transformation failures.
"""
import sys
import os
from pathlib import Path

# Add backend to path
backend_dir = Path(__file__).parent.parent
sys.path.insert(0, str(backend_dir))

import openpyxl
from app.services.bibbi.processors.liberty_processor import LibertyProcessor


def examine_file_structure(file_path: str):
    """Examine Liberty file structure"""
    print("=" * 80)
    print("LIBERTY FILE STRUCTURE ANALYSIS")
    print("=" * 80)

    wb = openpyxl.load_workbook(file_path, data_only=True)

    print(f"\nFile: {os.path.basename(file_path)}")
    print(f"Sheets: {wb.sheetnames}\n")

    for sheet_name in wb.sheetnames[:2]:  # First 2 sheets
        sheet = wb[sheet_name]
        print(f"\n--- SHEET: {sheet_name} ---")
        print(f"Dimensions: {sheet.dimensions}")

        # Show first 15 rows
        print("\nFirst 15 rows:")
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if i <= 15:
                # Truncate long values
                row_display = tuple(
                    str(cell)[:50] + "..." if cell and len(str(cell)) > 50 else cell
                    for cell in row
                )
                print(f"  Row {i}: {row_display}")

        print(f"\nTotal rows: {sum(1 for _ in sheet.iter_rows())}")

    wb.close()


def test_processor(file_path: str):
    """Test Liberty processor and capture errors"""
    print("\n" + "=" * 80)
    print("PROCESSOR TESTING")
    print("=" * 80)

    # Create processor instance
    reseller_id = "14b2a64e-013b-4c2d-9c42-379699b5823d"  # Liberty reseller ID
    batch_id = "test-debug-batch"

    processor = LibertyProcessor(reseller_id=reseller_id)

    print(f"\nProcessor: {processor.VENDOR_NAME}")
    print(f"Currency: {processor.CURRENCY}")
    print(f"Conversion rate: {processor.GBP_TO_EUR_RATE}")
    print(f"Column mapping: {processor.COLUMN_MAPPING}\n")

    # Step 1: Extract stores
    print("\n--- STEP 1: Extract Stores ---")
    try:
        stores = processor.extract_stores(file_path)
        print(f"✅ Extracted {len(stores)} stores:")
        for store in stores:
            print(f"   - {store}")
    except Exception as e:
        print(f"❌ Store extraction failed: {e}")
        import traceback
        traceback.print_exc()
        return

    # Step 2: Extract raw rows
    print("\n--- STEP 2: Extract Raw Rows ---")
    try:
        raw_rows = processor.extract_rows(file_path)
        print(f"✅ Extracted {len(raw_rows)} raw rows")

        # Show first 5 raw rows
        print("\nFirst 5 raw rows:")
        for i, row in enumerate(raw_rows[:5], start=1):
            print(f"  Row {i}: {row}")
    except Exception as e:
        print(f"❌ Row extraction failed: {e}")
        import traceback
        traceback.print_exc()
        return

    # Step 3: Transform rows
    print("\n--- STEP 3: Transform Rows ---")

    successful = []
    failed = []

    for row_num, raw_row in enumerate(raw_rows, start=2):
        try:
            transformed = processor.transform_row(raw_row, batch_id)
            if transformed:
                successful.append((row_num, transformed))
        except Exception as e:
            failed.append((row_num, raw_row, str(e)))

    print(f"\n✅ Successful: {len(successful)}")
    print(f"❌ Failed: {len(failed)}")

    # Show first 3 successful transformations
    if successful:
        print("\nFirst 3 successful transformations:")
        for row_num, data in successful[:3]:
            print(f"\n  Row {row_num}:")
            for key, value in data.items():
                print(f"    {key}: {value}")

    # Show first 10 failures
    if failed:
        print("\n\nFirst 10 transformation failures:")
        for row_num, raw_row, error in failed[:10]:
            print(f"\n  ❌ Row {row_num}:")
            print(f"     Raw data: {raw_row}")
            print(f"     Error: {error}")

    # Analyze error patterns
    if failed:
        print("\n\n--- ERROR PATTERN ANALYSIS ---")
        error_types = {}
        for _, _, error in failed:
            # Categorize error
            if "EAN" in error or "product_ean" in error:
                error_type = "EAN validation"
            elif "date" in error.lower():
                error_type = "Date parsing"
            elif "quantity" in error.lower():
                error_type = "Quantity parsing"
            elif "sales" in error.lower() or "value" in error.lower():
                error_type = "Sales value parsing"
            elif "store" in error.lower():
                error_type = "Store identification"
            else:
                error_type = "Other"

            error_types[error_type] = error_types.get(error_type, 0) + 1

        print("\nError distribution:")
        for error_type, count in sorted(error_types.items(), key=lambda x: x[1], reverse=True):
            print(f"  {error_type}: {count} rows")

    # Step 4: Run full process() method
    print("\n\n--- STEP 4: Run Full process() Method ---")
    try:
        result = processor.process(file_path, batch_id)
        print(f"\nProcessingResult:")
        print(f"  Vendor: {result.vendor}")
        print(f"  Total rows: {result.total_rows}")
        print(f"  Successful: {result.successful_rows}")
        print(f"  Failed: {result.failed_rows}")
        print(f"  Stores detected: {len(result.stores)}")
        print(f"  Transformed data records: {len(result.transformed_data)}")
        print(f"  Error records: {len(result.errors)}")

        # Show first 3 errors
        if result.errors:
            print("\nFirst 3 errors from ProcessingResult:")
            for error in result.errors[:3]:
                print(f"\n  Row {error.get('row_number')}:")
                print(f"    Error: {error.get('error')}")
                print(f"    Raw: {error.get('raw_data')}")

    except Exception as e:
        print(f"❌ Full process() failed: {e}")
        import traceback
        traceback.print_exc()


def main():
    file_path = "BIBBI/Example_docs/Liberty statistics/Continuity Supplier Size Report 31-08-2025.xlsx"

    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        print("\nLooking for Liberty files...")
        liberty_dir = "BIBBI/Example_docs/Liberty statistics"
        if os.path.exists(liberty_dir):
            files = [f for f in os.listdir(liberty_dir) if f.endswith('.xlsx')]
            print(f"Found {len(files)} files:")
            for f in files:
                print(f"  - {f}")

            if files:
                file_path = os.path.join(liberty_dir, files[0])
                print(f"\nUsing: {file_path}")
        else:
            print(f"Directory not found: {liberty_dir}")
            return

    # Run analysis
    examine_file_structure(file_path)
    test_processor(file_path)


if __name__ == "__main__":
    main()

"""
Generate sample test files for vendor processors
"""

import openpyxl
from pathlib import Path
from datetime import datetime


def generate_boxnox_test_file():
    """Generate a sample Boxnox Excel file for testing"""

    # Create workbook
    wb = openpyxl.Workbook()

    # Remove default sheet and create our target sheet
    wb.remove(wb.active)
    ws = wb.create_sheet("Sell Out by EAN")

    # Add headers
    headers = [
        "Product EAN",
        "Functional Name",
        "Sold Qty",
        "Sales Amount (EUR)",
        "Reseller",
        "Month",
        "Year"
    ]
    ws.append(headers)

    # Add sample data
    sample_data = [
        ["1234567890123", "Test Product Alpha", 10, 99.99, "Retailer A", 1, 2025],
        ["9876543210987", "Test Product Beta", 5, 49.99, "Retailer B", 1, 2025],
        ["5555555555555", "Test Product Gamma", 20, 199.99, "Retailer A", 1, 2025],
        ["1111111111111", "Test Product Delta", 15, 149.99, "Retailer C", 2, 2025],
        ["2222222222222", "Test Product Epsilon", 8, 79.99, "Retailer B", 2, 2025],
        ["3333333333333", "Test Product Zeta", 30, 299.99, "Retailer A", 2, 2025],
        ["4444444444444", "Test Product Eta", 12, 119.99, "Retailer C", 3, 2025],
        ["6666666666666", "Test Product Theta", 25, 249.99, "Retailer B", 3, 2025],
        ["7777777777777", "Test Product Iota", 18, 179.99, "Retailer A", 3, 2025],
        ["8888888888888", "Test Product Kappa", 22, 219.99, "Retailer C", 4, 2025],
    ]

    for row in sample_data:
        ws.append(row)

    # Style the header row
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    # Create test_data directory if it doesn't exist
    output_dir = Path(__file__).parent.parent / "test_data"
    output_dir.mkdir(exist_ok=True)

    # Save file
    output_file = output_dir / "test_boxnox.xlsx"
    wb.save(output_file)

    print(f"✅ Generated Boxnox test file: {output_file}")
    print(f"   - Sheet: 'Sell Out by EAN'")
    print(f"   - Rows: {len(sample_data)} data rows")
    print(f"   - Total records: {sum(row[2] for row in sample_data)} units sold")
    print(f"   - Total revenue: €{sum(row[3] for row in sample_data):.2f}")

    return output_file


def generate_boxnox_invalid_file():
    """Generate an invalid Boxnox file for testing error handling"""

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Sell Out by EAN")

    # Add headers
    headers = [
        "Product EAN",
        "Functional Name",
        "Sold Qty",
        "Sales Amount (EUR)",
        "Reseller",
        "Month",
        "Year"
    ]
    ws.append(headers)

    # Add invalid data to test error handling
    invalid_data = [
        ["12345", "Invalid EAN (too short)", 10, 99.99, "Retailer A", 1, 2025],  # Invalid EAN
        ["1234567890123", "Missing Month", 5, 49.99, "Retailer B", None, 2025],  # Missing month
        ["9876543210987", "Invalid Month", 20, 199.99, "Retailer A", 13, 2025],  # Month > 12
        ["5555555555555", "Invalid Year", 15, 149.99, "Retailer C", 2, 1999],  # Year < 2000
        ["abc123def4567", "Non-numeric EAN", 8, 79.99, "Retailer B", 2, 2025],  # Non-numeric EAN
    ]

    for row in invalid_data:
        ws.append(row)

    # Style the header row
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    output_dir = Path(__file__).parent.parent / "test_data"
    output_dir.mkdir(exist_ok=True)

    output_file = output_dir / "test_boxnox_invalid.xlsx"
    wb.save(output_file)

    print(f"✅ Generated invalid Boxnox test file: {output_file}")
    print(f"   - Contains {len(invalid_data)} invalid records for error testing")

    return output_file


def generate_all_test_files():
    """Generate all test files"""
    print("=" * 60)
    print("Generating Test Data Files")
    print("=" * 60)
    print()

    generate_boxnox_test_file()
    print()
    generate_boxnox_invalid_file()

    print()
    print("=" * 60)
    print("Test file generation complete!")
    print("=" * 60)


if __name__ == "__main__":
    generate_all_test_files()

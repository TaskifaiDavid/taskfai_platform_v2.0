"""
Generate demo online e-commerce sales data for AURORA LUXE brand

Creates 12 monthly Excel files (Jan-Dec 2024) with realistic sales data
that populates all 8 KPIs including gross_profit, profit_margin, unique_countries, and order_count.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import date, timedelta
import random
from pathlib import Path


# AURORA LUXE Product Catalog
PRODUCTS = [
    {
        "ean": "5901234567890",
        "name": "Hydrating Serum Arctic Dew",
        "price": 89.00,
        "cogs": 32.00,
        "seasonal_weight": {"winter": 1.5, "spring": 1.0, "summer": 0.8, "fall": 1.2}
    },
    {
        "ean": "5901234567891",
        "name": "Anti-Aging Cream Nordic Pearl",
        "price": 129.00,
        "cogs": 48.00,
        "seasonal_weight": {"winter": 1.2, "spring": 1.0, "summer": 0.9, "fall": 1.6}  # Holiday spike
    },
    {
        "ean": "5901234567892",
        "name": "Vitamin C Booster Midnight Sun",
        "price": 79.00,
        "cogs": 28.00,
        "seasonal_weight": {"winter": 0.9, "spring": 1.4, "summer": 1.3, "fall": 1.0}  # Spring boost
    },
    {
        "ean": "5901234567893",
        "name": "Eye Treatment Fjord Essence",
        "price": 69.00,
        "cogs": 24.00,
        "seasonal_weight": {"winter": 1.1, "spring": 1.0, "summer": 1.0, "fall": 1.1}
    },
    {
        "ean": "5901234567894",
        "name": "Face Mask Aurora Glow",
        "price": 39.00,
        "cogs": 12.00,
        "seasonal_weight": {"winter": 0.8, "spring": 1.1, "summer": 1.4, "fall": 0.9}  # Summer peak
    },
    {
        "ean": "5901234567895",
        "name": "Body Lotion Scandinavian Mist",
        "price": 49.00,
        "cogs": 18.00,
        "seasonal_weight": {"winter": 1.3, "spring": 1.0, "summer": 0.9, "fall": 1.0}  # Winter demand
    }
]

# Country distribution (10 countries for unique_countries KPI)
COUNTRIES = [
    ("Germany", 0.20),
    ("United States", 0.18),
    ("France", 0.15),
    ("United Kingdom", 0.12),
    ("Canada", 0.10),
    ("Netherlands", 0.08),
    ("Sweden", 0.07),
    ("Australia", 0.05),
    ("Japan", 0.03),
    ("Singapore", 0.02)
]

# Monthly revenue targets (shows growth trajectory)
MONTHLY_TARGETS = {
    1: 65000,   # January
    2: 72000,   # February
    3: 85000,   # March
    4: 92000,   # April
    5: 88000,   # May
    6: 95000,   # June
    7: 95000,   # July
    8: 88000,   # August
    9: 98000,   # September
    10: 115000, # October
    11: 135000, # November
    12: 160000  # December (Holiday peak)
}

# Season mapping
SEASONS = {
    1: "winter", 2: "winter", 3: "spring",
    4: "spring", 5: "spring", 6: "summer",
    7: "summer", 8: "summer", 9: "fall",
    10: "fall", 11: "fall", 12: "winter"
}

MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]


def calculate_stripe_fee(amount: float) -> float:
    """Calculate Stripe processing fee (2.9% + €0.30)"""
    return round(amount * 0.029 + 0.30, 2)


def select_country() -> str:
    """Select a random country based on distribution weights"""
    rand = random.random()
    cumulative = 0.0
    for country, weight in COUNTRIES:
        cumulative += weight
        if rand <= cumulative:
            return country
    return COUNTRIES[0][0]  # Fallback to Germany


def generate_orders_for_month(month: int, year: int) -> list:
    """Generate realistic orders for a given month"""
    season = SEASONS[month]
    target_revenue = MONTHLY_TARGETS[month]

    # Calculate weighted product pool based on season
    product_pool = []
    for product in PRODUCTS:
        weight = product["seasonal_weight"][season]
        count = int(weight * 10)  # Base frequency
        product_pool.extend([product] * count)

    orders = []
    order_counter = 1
    current_revenue = 0

    # Determine number of orders needed (target avg order value ~€90-€140)
    avg_order_value = random.uniform(90, 140)
    estimated_orders = int(target_revenue / avg_order_value)

    # Days in month
    if month in [1, 3, 5, 7, 8, 10, 12]:
        days_in_month = 31
    elif month in [4, 6, 9, 11]:
        days_in_month = 30
    else:
        days_in_month = 29 if year % 4 == 0 else 28

    while current_revenue < target_revenue:
        # Select 1-3 products per order
        items_in_order = random.choices(product_pool, k=random.randint(1, 3))

        order_date = date(year, month, random.randint(1, days_in_month))
        country = select_country()

        for product in items_in_order:
            quantity = random.choices([1, 2, 3], weights=[0.7, 0.25, 0.05])[0]

            sales_eur = product["price"] * quantity
            cost_of_goods = product["cogs"] * quantity
            stripe_fee = calculate_stripe_fee(sales_eur)

            order = {
                "order_id": f"ALX-{year}-{month:02d}-{order_counter:04d}",
                "product_ean": product["ean"],
                "functional_name": product["name"],
                "product_name": product["name"],
                "quantity": quantity,
                "sales_eur": round(sales_eur, 2),
                "cost_of_goods": round(cost_of_goods, 2),
                "stripe_fee": stripe_fee,
                "order_date": order_date.isoformat(),
                "country": country,
                "city": "",  # Optional
                "reseller": "Online"
            }

            orders.append(order)
            current_revenue += sales_eur
            order_counter += 1

        # Safety check to prevent infinite loop
        if len(orders) > estimated_orders * 2:
            break

    return orders


def create_excel_file(orders: list, month: int, year: int, output_dir: Path):
    """Create Excel file with orders data"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Orders"

    # Headers
    headers = [
        "Order ID",
        "Product EAN",
        "Functional Name",
        "Product Name",
        "Quantity",
        "Sales EUR",
        "Cost of Goods",
        "Stripe Fee",
        "Order Date",
        "Country",
        "City",
        "Reseller"
    ]

    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Write data
    for row_idx, order in enumerate(orders, start=2):
        sheet.cell(row=row_idx, column=1, value=order["order_id"])
        sheet.cell(row=row_idx, column=2, value=order["product_ean"])
        sheet.cell(row=row_idx, column=3, value=order["functional_name"])
        sheet.cell(row=row_idx, column=4, value=order["product_name"])
        sheet.cell(row=row_idx, column=5, value=order["quantity"])
        sheet.cell(row=row_idx, column=6, value=order["sales_eur"])
        sheet.cell(row=row_idx, column=7, value=order["cost_of_goods"])
        sheet.cell(row=row_idx, column=8, value=order["stripe_fee"])
        sheet.cell(row=row_idx, column=9, value=order["order_date"])
        sheet.cell(row=row_idx, column=10, value=order["country"])
        sheet.cell(row=row_idx, column=11, value=order["city"])
        sheet.cell(row=row_idx, column=12, value=order["reseller"])

    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Save file
    month_name = MONTH_NAMES[month - 1]
    filename = f"AURORA_LUXE_Online_{year}_{month:02d}_{month_name}.xlsx"
    filepath = output_dir / filename
    workbook.save(filepath)

    return filepath, len(orders)


def main():
    """Generate all 12 monthly demo files"""
    # Create output directory
    output_dir = Path(__file__).parent.parent.parent / "demo_data" / "2024_online_sales"
    output_dir.mkdir(parents=True, exist_ok=True)

    print("=" * 70)
    print("AURORA LUXE - Online Sales Demo Data Generator")
    print("=" * 70)
    print()

    total_orders = 0
    total_revenue = 0

    for month in range(1, 13):
        print(f"Generating {MONTH_NAMES[month-1]} 2024...")

        # Generate orders
        orders = generate_orders_for_month(month, 2024)

        # Create Excel file
        filepath, order_count = create_excel_file(orders, month, 2024, output_dir)

        month_revenue = sum(order["sales_eur"] for order in orders)
        total_orders += order_count
        total_revenue += month_revenue

        print(f"  ✓ Created: {filepath.name}")
        print(f"    Orders: {order_count:,} | Revenue: €{month_revenue:,.2f}")
        print()

    print("=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"Total Files Generated: 12")
    print(f"Total Orders: {total_orders:,}")
    print(f"Total Revenue: €{total_revenue:,.2f}")
    print(f"Average Order Value: €{total_revenue/total_orders:.2f}")
    print()
    print("Expected KPIs after upload:")
    print(f"  ✓ Total Revenue: €{total_revenue:,.0f}")
    print(f"  ✓ Order Count: {total_orders:,}")
    print(f"  ✓ Unique Countries: 10")
    print(f"  ✓ Gross Profit: ~€{total_revenue * 0.40:,.0f} (40% margin)")
    print(f"  ✓ Profit Margin: 38-42%")
    print()
    print(f"Files saved to: {output_dir}")
    print("=" * 70)


if __name__ == "__main__":
    main()

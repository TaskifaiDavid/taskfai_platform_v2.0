"""
Excel file processing utilities for vendor data

This module contains common Excel parsing and extraction utilities
extracted from vendor processors to eliminate code duplication.
"""

from typing import List, Dict, Any, Optional
from openpyxl.worksheet.worksheet import Worksheet
import openpyxl


def extract_rows_from_sheet(
    sheet: Worksheet,
    header_row: int = 1,
    min_data_row: int = 2,
    skip_empty: bool = True
) -> List[Dict[str, Any]]:
    """
    Extract rows from Excel worksheet as list of dictionaries

    Standard pattern used across all vendor processors:
    - First row contains headers
    - Data starts from second row
    - Empty rows are skipped
    - Each row is converted to dictionary with headers as keys

    Args:
        sheet: Excel worksheet object
        header_row: Row number containing headers (default: 1)
        min_data_row: First row containing data (default: 2)
        skip_empty: Skip rows where all values are None/empty (default: True)

    Returns:
        List of dictionaries where keys are column headers

    Examples:
        >>> # Excel sheet:
        >>> # Row 1: | EAN         | Product | Qty |
        >>> # Row 2: | 1234567890123 | Test   | 10  |
        >>> sheet = workbook["Sheet1"]
        >>> rows = extract_rows_from_sheet(sheet)
        >>> rows[0]
        {'EAN': '1234567890123', 'Product': 'Test', 'Qty': 10}
    """
    # Extract headers from first row
    headers = []
    for cell in sheet[header_row]:
        if cell.value:
            headers.append(str(cell.value).strip())

    # Handle case where sheet has no headers
    if not headers:
        raise ValueError("No headers found in worksheet")

    # Extract data rows
    rows = []
    for row in sheet.iter_rows(min_row=min_data_row, values_only=True):
        # Skip empty rows if requested
        if skip_empty and not any(row):
            continue

        # Build dictionary from row
        row_dict = {}
        for idx, header in enumerate(headers):
            if idx < len(row):
                row_dict[header] = row[idx]

        rows.append(row_dict)

    return rows


def find_sheet_by_name(
    workbook: openpyxl.Workbook,
    sheet_name: str,
    fallback_to_first: bool = False
) -> Worksheet:
    """
    Find worksheet by name with optional fallback

    Args:
        workbook: Excel workbook object
        sheet_name: Name of sheet to find
        fallback_to_first: If True, returns first sheet when named sheet not found

    Returns:
        Worksheet object

    Raises:
        ValueError: If sheet not found and fallback_to_first=False

    Examples:
        >>> workbook = openpyxl.load_workbook("file.xlsx")
        >>> sheet = find_sheet_by_name(workbook, "Sell Out")
        >>> # Or with fallback:
        >>> sheet = find_sheet_by_name(workbook, "Data", fallback_to_first=True)
    """
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]

    if fallback_to_first:
        return workbook[workbook.sheetnames[0]]

    raise ValueError(
        f"Sheet '{sheet_name}' not found in workbook. "
        f"Available sheets: {', '.join(workbook.sheetnames)}"
    )


def get_sheet_headers(sheet: Worksheet, header_row: int = 1) -> List[str]:
    """
    Extract header names from worksheet

    Args:
        sheet: Excel worksheet object
        header_row: Row number containing headers (default: 1)

    Returns:
        List of header names (cleaned and stripped)

    Examples:
        >>> sheet = workbook["Data"]
        >>> headers = get_sheet_headers(sheet)
        >>> headers
        ['Product EAN', 'Functional Name', 'Quantity', 'Sales Amount']
    """
    headers = []
    for cell in sheet[header_row]:
        if cell.value:
            headers.append(str(cell.value).strip())

    return headers


def validate_required_headers(
    sheet: Worksheet,
    required_headers: List[str],
    header_row: int = 1
) -> None:
    """
    Validate that worksheet contains all required headers

    Args:
        sheet: Excel worksheet object
        required_headers: List of header names that must be present
        header_row: Row number containing headers (default: 1)

    Raises:
        ValueError: If any required headers are missing

    Examples:
        >>> required = ["Product EAN", "Quantity", "Sales Amount"]
        >>> validate_required_headers(sheet, required)
        >>> # Raises ValueError if any column is missing
    """
    actual_headers = get_sheet_headers(sheet, header_row)

    missing_headers = []
    for required in required_headers:
        if required not in actual_headers:
            missing_headers.append(required)

    if missing_headers:
        raise ValueError(
            f"Missing required columns: {', '.join(missing_headers)}. "
            f"Found columns: {', '.join(actual_headers)}"
        )


def count_data_rows(
    sheet: Worksheet,
    min_row: int = 2,
    skip_empty: bool = True
) -> int:
    """
    Count number of data rows in worksheet

    Args:
        sheet: Excel worksheet object
        min_row: First row to count (default: 2, assuming row 1 is headers)
        skip_empty: Skip rows where all values are None/empty (default: True)

    Returns:
        Number of data rows

    Examples:
        >>> count = count_data_rows(sheet)
        >>> print(f"Found {count} rows of data")
    """
    count = 0
    for row in sheet.iter_rows(min_row=min_row, values_only=True):
        if skip_empty and not any(row):
            continue
        count += 1

    return count


def safe_load_workbook(
    file_path: str,
    data_only: bool = True,
    read_only: bool = False
) -> openpyxl.Workbook:
    """
    Safely load Excel workbook with error handling

    Args:
        file_path: Path to Excel file
        data_only: If True, reads cell values instead of formulas (default: True)
        read_only: If True, opens in read-only mode for better performance

    Returns:
        Workbook object

    Raises:
        ValueError: If file cannot be loaded or is not a valid Excel file

    Examples:
        >>> workbook = safe_load_workbook("/path/to/file.xlsx")
        >>> # Process workbook...
        >>> workbook.close()
    """
    try:
        return openpyxl.load_workbook(
            file_path,
            data_only=data_only,
            read_only=read_only
        )
    except FileNotFoundError:
        raise ValueError(f"Excel file not found: {file_path}")
    except Exception as e:
        raise ValueError(f"Failed to load Excel file: {str(e)}")

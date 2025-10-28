"""
Base Vendor Processor

Abstract base class for vendor data processors.
This base class provides common utilities shared across all vendor processors.

NOTE: For BIBBI-specific processors, use app.services.bibbi.processors.base.BibbiBseProcessor
which extends this with BIBBI tenant-specific business logic.
"""

from abc import ABC, abstractmethod
from datetime import datetime
from typing import List, Dict, Any, Optional
from decimal import Decimal, InvalidOperation
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.validation import validate_ean, to_int, to_float, validate_month, validate_year
from app.utils.excel import (
    extract_rows_from_sheet,
    get_sheet_headers,
    safe_load_workbook,
    find_sheet_by_name
)


class VendorProcessor(ABC):
    """
    Abstract base class for vendor data processors

    Each vendor processor must implement:
    - process(): Main processing pipeline
    - get_vendor_name(): Return vendor identifier

    Common utilities provided:
    - _load_workbook(): Load Excel files safely
    - _get_sheet_headers(): Extract column headers
    - _extract_rows(): Parse rows as dictionaries
    - _validate_ean(): Validate EAN codes
    - _to_int(), _to_float(), _to_decimal(): Type conversions
    - _validate_month(), _validate_year(): Date validation
    """

    # Currency conversion rates (approximate - should be configurable)
    CURRENCY_RATES = {
        "EUR": 1.0,      # Base currency
        "GBP": 1.17,     # GBP to EUR
        "USD": 0.92,     # USD to EUR
        "PLN": 0.23,     # Polish Zloty to EUR
        "ZAR": 0.05,     # South African Rand to EUR
    }

    def __init__(self, reseller_id: Optional[str] = None):
        """
        Initialize processor

        Args:
            reseller_id: Optional reseller identifier
        """
        self.reseller_id = reseller_id

    @abstractmethod
    def process(
        self,
        file_path: str,
        user_id: str,
        batch_id: str
    ) -> Dict[str, Any]:
        """
        Process vendor file

        Args:
            file_path: Path to Excel file
            user_id: User identifier
            batch_id: Batch identifier

        Returns:
            Processing result with statistics:
            {
                "total_rows": int,
                "successful_rows": int,
                "failed_rows": int,
                "transformed_data": List[Dict],
                "errors": List[Dict]
            }
        """
        pass

    @abstractmethod
    def get_vendor_name(self) -> str:
        """Return vendor identifier (e.g., 'boxnox', 'galilu')"""
        pass

    # Excel Utilities

    def _load_workbook(
        self,
        file_path: str,
        read_only: bool = False
    ) -> openpyxl.Workbook:
        """
        Load Excel workbook safely

        Uses shared utility: app.utils.excel.safe_load_workbook

        Args:
            file_path: Path to Excel file
            read_only: If True, opens in read-only mode

        Returns:
            Workbook object
        """
        return safe_load_workbook(file_path, data_only=True, read_only=read_only)

    def _find_sheet(
        self,
        workbook: openpyxl.Workbook,
        sheet_name: str,
        fallback_to_first: bool = False
    ) -> Worksheet:
        """
        Find worksheet by name with optional fallback

        Uses shared utility: app.utils.excel.find_sheet_by_name

        Args:
            workbook: Excel workbook
            sheet_name: Name of sheet to find
            fallback_to_first: If True, returns first sheet when named sheet not found

        Returns:
            Worksheet object
        """
        return find_sheet_by_name(workbook, sheet_name, fallback_to_first)

    def _get_sheet_headers(self, sheet: Worksheet) -> List[str]:
        """
        Extract column headers from first row

        Uses shared utility: app.utils.excel.get_sheet_headers

        Args:
            sheet: Excel worksheet

        Returns:
            List of header names
        """
        return get_sheet_headers(sheet, header_row=1)

    def _extract_rows(
        self,
        sheet: Worksheet,
        min_row: int = 2
    ) -> List[Dict[str, Any]]:
        """
        Extract rows from worksheet as dictionaries

        Uses shared utility: app.utils.excel.extract_rows_from_sheet

        Args:
            sheet: Excel worksheet
            min_row: First row containing data (default: 2, assumes row 1 is headers)

        Returns:
            List of row dictionaries
        """
        return extract_rows_from_sheet(sheet, header_row=1, min_data_row=min_row)

    # Validation Utilities

    def _validate_ean(
        self,
        value: Any,
        required: bool = True,
        strict: bool = True
    ) -> Optional[str]:
        """
        Validate and normalize EAN code

        Uses shared utility: app.utils.validation.validate_ean

        Args:
            value: EAN value from Excel
            required: If False, returns None for empty values
            strict: If False, returns None for invalid EAN instead of raising

        Returns:
            Normalized 13-digit EAN string or None

        Raises:
            ValueError: If EAN is invalid and required=True and strict=True
        """
        return validate_ean(value, required=required, strict=strict)

    def _validate_month(self, value: Any) -> int:
        """
        Validate month value (1-12)

        Uses shared utility: app.utils.validation.validate_month

        Args:
            value: Raw month value

        Returns:
            Validated month as integer (1-12)
        """
        return validate_month(value)

    def _validate_year(
        self,
        value: Any,
        min_year: int = 2000,
        max_year: int = 2100
    ) -> int:
        """
        Validate year value

        Uses shared utility: app.utils.validation.validate_year

        Args:
            value: Raw year value
            min_year: Minimum acceptable year
            max_year: Maximum acceptable year

        Returns:
            Validated year as integer
        """
        return validate_year(value, min_year, max_year)

    # Type Conversion Utilities

    def _to_int(self, value: Any, field_name: str) -> int:
        """
        Convert value to integer safely

        Uses shared utility: app.utils.validation.to_int

        Args:
            value: Value to convert
            field_name: Field name for error messages

        Returns:
            Integer value

        Raises:
            ValueError: If conversion fails
        """
        return to_int(value, field_name)

    def _to_float(
        self,
        value: Any,
        field_name: str,
        allow_none: bool = False,
        default: float = 0.0
    ) -> float:
        """
        Convert value to float safely

        Uses shared utility: app.utils.validation.to_float

        Args:
            value: Value to convert
            field_name: Field name for error messages
            allow_none: If True, returns default for None/empty values
            default: Default value when allow_none=True

        Returns:
            Float value

        Raises:
            ValueError: If conversion fails and allow_none=False
        """
        return to_float(value, field_name, allow_none=allow_none, default=default)

    def _to_decimal(self, value: Any, field_name: str) -> Decimal:
        """
        Convert value to Decimal safely

        Args:
            value: Value to convert
            field_name: Field name for error messages

        Returns:
            Decimal value

        Raises:
            ValueError: If conversion fails
        """
        try:
            float_val = self._to_float(value, field_name)
            return Decimal(str(float_val))
        except (ValueError, InvalidOperation):
            raise ValueError(f"Invalid decimal for {field_name}: {value}")

    # Currency Conversion

    def _convert_currency(
        self,
        amount: float,
        from_currency: str,
        to_currency: str = "EUR"
    ) -> float:
        """
        Convert amount between currencies

        Args:
            amount: Amount in source currency
            from_currency: Source currency code (EUR, GBP, PLN, ZAR, USD)
            to_currency: Target currency code (default: EUR)

        Returns:
            Amount in target currency

        Raises:
            ValueError: If currency is unknown
        """
        if from_currency == to_currency:
            return amount

        from_rate = self.CURRENCY_RATES.get(from_currency)
        to_rate = self.CURRENCY_RATES.get(to_currency)

        if from_rate is None:
            raise ValueError(f"Unknown currency: {from_currency}")
        if to_rate is None:
            raise ValueError(f"Unknown currency: {to_currency}")

        # Convert to EUR first, then to target currency
        eur_amount = amount * from_rate
        return round(eur_amount / to_rate, 2)

    # Date Utilities

    def _validate_date(self, value: Any) -> datetime:
        """
        Validate and parse date value

        Args:
            value: Date value (datetime object or string)

        Returns:
            datetime object

        Raises:
            ValueError: If date is invalid
        """
        if isinstance(value, datetime):
            return value

        if isinstance(value, str):
            # Try common date formats
            for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]:
                try:
                    return datetime.strptime(value, fmt)
                except ValueError:
                    continue

            raise ValueError(f"Invalid date format: {value}")

        raise ValueError(f"Invalid date type: {type(value)}")

    def _calculate_quarter(self, month: int) -> int:
        """
        Calculate quarter from month (1-4)

        Args:
            month: Month number (1-12)

        Returns:
            Quarter number (1-4)
        """
        return (month - 1) // 3 + 1

    # Helper Methods

    def _create_error_dict(
        self,
        row_number: int,
        error: Exception,
        raw_data: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        Create standardized error dictionary

        Args:
            row_number: Row number where error occurred
            error: Exception that was raised
            raw_data: Raw row data that caused error

        Returns:
            Error dictionary for reporting
        """
        return {
            "row_number": row_number,
            "error": str(error),
            "raw_data": raw_data
        }

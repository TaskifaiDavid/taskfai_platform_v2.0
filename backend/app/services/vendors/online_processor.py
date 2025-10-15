"""
Online/E-commerce vendor data processor

Processes online sales data and inserts into ecommerce_orders table.
This enables all 8 KPIs including gross_profit, profit_margin, unique_countries, and order_count.
"""

from typing import Dict, List, Any, Optional
from datetime import datetime, date
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class OnlineProcessor:
    """Process online e-commerce Excel files"""

    TARGET_SHEET = "Orders"

    COLUMN_MAPPING = {
        "Order ID": "order_id_ref",
        "Product EAN": "product_ean",
        "Functional Name": "functional_name",
        "Product Name": "product_name",
        "Quantity": "quantity",
        "Sales EUR": "sales_eur",
        "Cost of Goods": "cost_of_goods",
        "Stripe Fee": "stripe_fee",
        "Order Date": "order_date",
        "Country": "country",
        "City": "city",
        "Reseller": "reseller"
    }

    def process(
        self,
        file_path: str,
        user_id: str,
        batch_id: str
    ) -> Dict[str, Any]:
        """
        Process online e-commerce file

        Args:
            file_path: Path to Excel file
            user_id: User identifier
            batch_id: Batch identifier

        Returns:
            Processing result with statistics
        """
        try:
            # Load workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            # Check if target sheet exists
            if self.TARGET_SHEET not in workbook.sheetnames:
                # Try first sheet if Orders sheet not found
                if workbook.sheetnames:
                    sheet = workbook[workbook.sheetnames[0]]
                else:
                    raise ValueError("No sheets found in workbook")
            else:
                sheet = workbook[self.TARGET_SHEET]

            # Extract data
            raw_rows = self._extract_rows(sheet)

            # Transform data
            transformed_rows = []
            errors = []

            for row_num, raw_row in enumerate(raw_rows, start=2):  # Start from 2 (after header)
                try:
                    transformed = self._transform_row(raw_row, user_id, batch_id)
                    if transformed:
                        transformed_rows.append(transformed)
                except Exception as e:
                    errors.append({
                        "row_number": row_num,
                        "error": str(e),
                        "raw_data": raw_row
                    })

            workbook.close()

            return {
                "total_rows": len(raw_rows),
                "successful_rows": len(transformed_rows),
                "failed_rows": len(errors),
                "transformed_data": transformed_rows,
                "errors": errors
            }

        except Exception as e:
            raise Exception(f"Failed to process online e-commerce file: {str(e)}")

    def _extract_rows(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """
        Extract rows from worksheet

        Args:
            sheet: Excel worksheet

        Returns:
            List of row dictionaries
        """
        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())

        # Extract data rows
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Skip empty rows
            if not any(row):
                continue

            row_dict = {}
            for idx, header in enumerate(headers):
                if idx < len(row):
                    row_dict[header] = row[idx]

            rows.append(row_dict)

        return rows

    def _transform_row(
        self,
        raw_row: Dict[str, Any],
        user_id: str,
        batch_id: str
    ) -> Optional[Dict[str, Any]]:
        """
        Transform raw row to ecommerce_orders format

        Args:
            raw_row: Raw row data
            user_id: User identifier
            batch_id: Batch identifier

        Returns:
            Transformed row or None if invalid
        """
        # Map columns
        transformed = {
            "user_id": user_id,
            "upload_batch_id": batch_id
        }

        for source_col, target_col in self.COLUMN_MAPPING.items():
            value = raw_row.get(source_col)

            # Handle required fields with validation
            if target_col == "product_ean":
                # EAN can be null for some products
                transformed[target_col] = self._validate_ean(value) if value else None
            elif target_col == "functional_name":
                if not value:
                    raise ValueError("Missing required field: Functional Name")
                transformed[target_col] = str(value).strip()
            elif target_col == "product_name":
                # Product name optional, defaults to functional name
                transformed[target_col] = str(value).strip() if value else transformed.get("functional_name")
            elif target_col == "quantity":
                if not value:
                    raise ValueError("Missing required field: Quantity")
                transformed[target_col] = self._to_int(value, "Quantity")
            elif target_col == "sales_eur":
                if not value:
                    raise ValueError("Missing required field: Sales EUR")
                transformed[target_col] = self._to_float(value, "Sales EUR")
            elif target_col == "cost_of_goods":
                # COGS is optional but recommended for profit calculations
                transformed[target_col] = self._to_float(value, "Cost of Goods") if value else None
            elif target_col == "stripe_fee":
                # Stripe fee is optional but recommended for profit calculations
                transformed[target_col] = self._to_float(value, "Stripe Fee") if value else None
            elif target_col == "order_date":
                if not value:
                    raise ValueError("Missing required field: Order Date")
                transformed[target_col] = self._parse_date(value)
            elif target_col == "country":
                # Country is important for unique_countries KPI
                transformed[target_col] = str(value).strip() if value else None
            elif target_col == "city":
                transformed[target_col] = str(value).strip() if value else None
            elif target_col == "reseller":
                # Default to "Online" if not specified
                transformed[target_col] = str(value).strip() if value else "Online"
            elif target_col == "order_id_ref":
                # Store reference order ID if provided
                transformed[target_col] = str(value).strip() if value else None

        return transformed

    def _validate_ean(self, value: Any) -> Optional[str]:
        """Validate EAN code"""
        if not value:
            return None

        ean_str = str(value).strip()

        # Remove any decimal points (in case Excel formatted as number)
        if '.' in ean_str:
            ean_str = ean_str.split('.')[0]

        # Check length (13 digits for EAN-13)
        if len(ean_str) == 13 and ean_str.isdigit():
            return ean_str

        # Return None for invalid EANs (don't raise error)
        return None

    def _parse_date(self, value: Any) -> str:
        """Parse date value to ISO format"""
        if isinstance(value, date):
            return value.isoformat()
        elif isinstance(value, datetime):
            return value.date().isoformat()
        elif isinstance(value, str):
            # Try to parse string date
            try:
                # Try ISO format first
                parsed = datetime.fromisoformat(value.strip())
                return parsed.date().isoformat()
            except:
                try:
                    # Try common formats
                    for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]:
                        try:
                            parsed = datetime.strptime(value.strip(), fmt)
                            return parsed.date().isoformat()
                        except:
                            continue
                    raise ValueError(f"Unable to parse date: {value}")
                except:
                    raise ValueError(f"Invalid date format: {value}")
        else:
            raise ValueError(f"Invalid date type: {type(value)}")

    def _to_int(self, value: Any, field_name: str) -> int:
        """Convert value to integer"""
        if value is None:
            raise ValueError(f"{field_name} cannot be None")

        try:
            return int(float(value))
        except (ValueError, TypeError):
            raise ValueError(f"Invalid integer for {field_name}: {value}")

    def _to_float(self, value: Any, field_name: str) -> Optional[float]:
        """Convert value to float"""
        if value is None or value == "":
            return None

        try:
            return float(value)
        except (ValueError, TypeError):
            raise ValueError(f"Invalid float for {field_name}: {value}")

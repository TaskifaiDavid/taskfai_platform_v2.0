"""
CDLC vendor data processor (Multi-month aggregation support)
"""

from typing import Dict, List, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import re


class CDLCProcessor:
    """Process CDLC Excel files with multi-month aggregation"""

    TARGET_SHEET = "Sheet1"

    COLUMN_MAPPING = {
        "Product": "functional_name",
        "EAN": "product_ean",
        "Total": "sales_eur",
        "Reseller": "reseller"
    }

    def process(
        self,
        file_path: str,
        user_id: str,
        batch_id: str
    ) -> Dict[str, Any]:
        """
        Process CDLC file with multi-month data

        Args:
            file_path: Path to Excel file
            user_id: User identifier
            batch_id: Batch identifier

        Returns:
            Processing result with statistics
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            if self.TARGET_SHEET not in workbook.sheetnames:
                sheet = workbook[workbook.sheetnames[0]]
            else:
                sheet = workbook[self.TARGET_SHEET]

            raw_rows = self._extract_rows(sheet)

            transformed_rows = []
            errors = []

            for row_num, raw_row in enumerate(raw_rows, start=2):
                try:
                    # CDLC may have multiple month columns
                    transformed_list = self._transform_row(raw_row, user_id, batch_id)
                    if transformed_list:
                        transformed_rows.extend(transformed_list)
                except Exception as e:
                    errors.append({
                        "row_number": row_num,
                        "error": str(e),
                        "raw_data": raw_row
                    })

            workbook.close()

            return {
                "total_rows": len(raw_rows),
                "successful_rows": len(transformed_rows),
                "failed_rows": len(errors),
                "transformed_data": transformed_rows,
                "errors": errors
            }

        except Exception as e:
            raise Exception(f"Failed to process CDLC file: {str(e)}")

    def _extract_rows(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """Extract rows from worksheet"""
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())

        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            row_dict = {}
            for idx, header in enumerate(headers):
                if idx < len(row):
                    row_dict[header] = row[idx]

            rows.append(row_dict)

        return rows

    def _transform_row(
        self,
        raw_row: Dict[str, Any],
        user_id: str,
        batch_id: str
    ) -> List[Dict[str, Any]]:
        """
        Transform raw row - may produce multiple rows for multi-month data

        CDLC files may have columns like "Jan 2024", "Feb 2024", etc.
        """
        base_data = {
            "user_id": user_id,
            "batch_id": batch_id,
            "vendor": "cdlc",
            "currency": "EUR"
        }

        # Extract base fields
        for source_col, target_col in self.COLUMN_MAPPING.items():
            value = raw_row.get(source_col)

            if source_col in ["Product"] and not value:
                raise ValueError(f"Missing required field: {source_col}")

            if target_col == "product_ean":
                base_data[target_col] = self._validate_ean(value) if value else None
            elif target_col == "functional_name":
                base_data[target_col] = str(value)
            elif target_col == "reseller":
                base_data[target_col] = str(value) if value else "CDLC"
            elif target_col == "sales_eur":
                # This might be a total column, handle separately
                pass

        # Check for month columns (e.g., "Jan 2024", "Feb 2024", etc.)
        month_pattern = re.compile(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{4})')

        results = []
        for col_name, value in raw_row.items():
            match = month_pattern.match(col_name)
            if match and value:
                month_str = match.group(1)
                year = int(match.group(2))
                month = self._month_name_to_number(month_str)

                row_data = base_data.copy()
                row_data["month"] = month
                row_data["year"] = year
                row_data["sales_eur"] = self._to_float(value, col_name)
                row_data["quantity"] = 1  # CDLC may not have quantity
                row_data["created_at"] = datetime.utcnow().isoformat()

                results.append(row_data)

        # If no month columns found, use Total column
        if not results and "Total" in raw_row:
            row_data = base_data.copy()
            # Use current date for month/year if not specified
            current_date = datetime.utcnow()
            row_data["month"] = current_date.month
            row_data["year"] = current_date.year
            row_data["sales_eur"] = self._to_float(raw_row["Total"], "Total")
            row_data["quantity"] = 1
            row_data["created_at"] = current_date.isoformat()
            results.append(row_data)

        return results

    def _month_name_to_number(self, month_name: str) -> int:
        """Convert month name to number"""
        months = {
            "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4,
            "May": 5, "Jun": 6, "Jul": 7, "Aug": 8,
            "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12
        }
        return months.get(month_name, 1)

    def _validate_ean(self, value: Any) -> Optional[str]:
        """Validate EAN code"""
        if not value:
            return None

        ean_str = str(value).strip()
        if '.' in ean_str:
            ean_str = ean_str.split('.')[0]

        if len(ean_str) != 13 or not ean_str.isdigit():
            return None

        return ean_str

    def _to_float(self, value: Any, field_name: str) -> float:
        """Convert value to float"""
        if value is None or value == "":
            return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            raise ValueError(f"Invalid float for {field_name}: {value}")

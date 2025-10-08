"""
Continuity vendor data processor (UK - Special fields handling)
"""

from typing import Dict, List, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class ContinuityProcessor:
    """Process Continuity Excel files with special field handling"""

    TARGET_SHEET = "Sheet1"

    COLUMN_MAPPING = {
        "EAN": "product_ean",
        "Product": "functional_name",
        "Units": "quantity",
        "Value": "sales_eur",
        "Period": "period",  # Special: May contain "Q1 2024" or "Jan-Mar 2024"
        "Channel": "channel"  # Special: Distribution channel
    }

    def process(
        self,
        file_path: str,
        user_id: str,
        batch_id: str
    ) -> Dict[str, Any]:
        """
        Process Continuity file with special field handling

        Args:
            file_path: Path to Excel file
            user_id: User identifier
            batch_id: Batch identifier

        Returns:
            Processing result with statistics
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            if self.TARGET_SHEET not in workbook.sheetnames:
                sheet = workbook[workbook.sheetnames[0]]
            else:
                sheet = workbook[self.TARGET_SHEET]

            raw_rows = self._extract_rows(sheet)

            transformed_rows = []
            errors = []

            for row_num, raw_row in enumerate(raw_rows, start=2):
                try:
                    transformed = self._transform_row(raw_row, user_id, batch_id)
                    if transformed:
                        transformed_rows.append(transformed)
                except Exception as e:
                    errors.append({
                        "row_number": row_num,
                        "error": str(e),
                        "raw_data": raw_row
                    })

            workbook.close()

            return {
                "total_rows": len(raw_rows),
                "successful_rows": len(transformed_rows),
                "failed_rows": len(errors),
                "transformed_data": transformed_rows,
                "errors": errors
            }

        except Exception as e:
            raise Exception(f"Failed to process Continuity file: {str(e)}")

    def _extract_rows(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """Extract rows from worksheet"""
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())

        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            row_dict = {}
            for idx, header in enumerate(headers):
                if idx < len(row):
                    row_dict[header] = row[idx]

            rows.append(row_dict)

        return rows

    def _transform_row(
        self,
        raw_row: Dict[str, Any],
        user_id: str,
        batch_id: str
    ) -> Optional[Dict[str, Any]]:
        """Transform raw row with special field handling"""
        transformed = {
            "user_id": user_id,
            "batch_id": batch_id,
            "vendor": "continuity",
            "currency": "EUR",
            "reseller": "Continuity Suppliers"
        }

        for source_col, target_col in self.COLUMN_MAPPING.items():
            value = raw_row.get(source_col)

            if source_col in ["EAN", "Product", "Units"] and not value:
                raise ValueError(f"Missing required field: {source_col}")

            if target_col == "product_ean":
                transformed[target_col] = self._validate_ean(value)
            elif target_col == "quantity":
                transformed[target_col] = self._to_int(value, "Units")
            elif target_col == "sales_eur":
                transformed[target_col] = self._to_float(value, "Value")
            elif target_col == "period":
                # Parse period (e.g., "Q1 2024" or "Jan-Mar 2024" or "Jan 2024")
                month, year = self._parse_period(value)
                transformed["month"] = month
                transformed["year"] = year
                transformed["period_text"] = str(value) if value else None
            elif target_col == "channel":
                # Store channel information (e.g., "Retail", "Wholesale", "Online")
                transformed["channel"] = str(value) if value else "Retail"
            elif target_col == "functional_name":
                transformed[target_col] = str(value)
            else:
                transformed[target_col] = str(value) if value is not None else None

        transformed["created_at"] = datetime.utcnow().isoformat()

        return transformed

    def _parse_period(self, value: Any) -> tuple[int, int]:
        """
        Parse period string to extract month and year

        Handles formats:
        - "Q1 2024" -> (1, 2024)
        - "Jan 2024" -> (1, 2024)
        - "Jan-Mar 2024" -> (1, 2024) # Use first month
        - "2024-01" -> (1, 2024)
        """
        if not value:
            # Use current date
            now = datetime.utcnow()
            return now.month, now.year

        period_str = str(value).strip()

        # Handle quarter format "Q1 2024"
        if period_str.startswith('Q'):
            parts = period_str.split()
            if len(parts) >= 2:
                quarter = int(parts[0][1])  # Extract number from "Q1"
                year = int(parts[1])
                month = (quarter - 1) * 3 + 1  # Q1->1, Q2->4, Q3->7, Q4->10
                return month, year

        # Handle month name formats
        month_names = {
            "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
            "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12
        }

        for month_name, month_num in month_names.items():
            if month_name in period_str.lower():
                # Extract year
                parts = period_str.split()
                year = int(parts[-1]) if parts else datetime.utcnow().year
                return month_num, year

        # Handle YYYY-MM format
        if '-' in period_str and len(period_str.split('-')) == 2:
            year_str, month_str = period_str.split('-')
            return int(month_str), int(year_str)

        # Default to current month/year
        now = datetime.utcnow()
        return now.month, now.year

    def _validate_ean(self, value: Any) -> str:
        """Validate EAN code"""
        if not value:
            raise ValueError("EAN cannot be empty")

        ean_str = str(value).strip()
        if '.' in ean_str:
            ean_str = ean_str.split('.')[0]

        if len(ean_str) != 13 or not ean_str.isdigit():
            raise ValueError(f"Invalid EAN format: {ean_str}")

        return ean_str

    def _to_int(self, value: Any, field_name: str) -> int:
        """Convert value to integer"""
        if value is None:
            raise ValueError(f"{field_name} cannot be None")
        try:
            return int(float(value))
        except (ValueError, TypeError):
            raise ValueError(f"Invalid integer for {field_name}: {value}")

    def _to_float(self, value: Any, field_name: str) -> float:
        """Convert value to float"""
        if value is None or value == "":
            return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            raise ValueError(f"Invalid float for {field_name}: {value}")

"""
Demo vendor data processor for testing and demonstrations
"""

from typing import Dict, List, Any
from datetime import datetime
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import uuid


class DemoProcessor:
    """Process Demo Excel files with simple structure"""

    TARGET_SHEET = "Sheet1"

    COLUMN_MAPPING = {
        "Brand": "product_ean",
        "BrandName": "functional_name",
        "Month": "month",
        "Information": "description"
    }

    def process(
        self,
        file_path: str,
        user_id: str,
        batch_id: str
    ) -> Dict[str, Any]:
        """
        Process Demo file

        Args:
            file_path: Path to Excel file
            user_id: User identifier
            batch_id: Batch identifier

        Returns:
            Processing result with statistics
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            if self.TARGET_SHEET not in workbook.sheetnames:
                sheet = workbook[workbook.sheetnames[0]]
            else:
                sheet = workbook[self.TARGET_SHEET]

            raw_rows = self._extract_rows(sheet)

            transformed_rows = []
            errors = []

            for row_num, raw_row in enumerate(raw_rows, start=2):
                try:
                    transformed = self._transform_row(raw_row, user_id, batch_id)
                    if transformed:
                        transformed_rows.append(transformed)
                except Exception as e:
                    errors.append({
                        "row_number": row_num,
                        "error": str(e),
                        "raw_data": raw_row
                    })

            workbook.close()

            return {
                "total_rows": len(raw_rows),
                "successful_rows": len(transformed_rows),
                "failed_rows": len(errors),
                "transformed_data": transformed_rows,
                "errors": errors
            }

        except Exception as e:
            raise Exception(f"Failed to process Demo file: {str(e)}")

    def _extract_rows(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """Extract data rows from worksheet"""
        rows = []
        headers = []

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                # First row is headers
                headers = [str(cell).strip() if cell else f"Column_{i}" for i, cell in enumerate(row)]
                continue

            # Skip empty rows
            if not any(row):
                continue

            # Create dict from row
            row_dict = {}
            for header, value in zip(headers, row):
                row_dict[header] = value

            rows.append(row_dict)

        return rows

    def _transform_row(
        self,
        raw_row: Dict[str, Any],
        user_id: str,
        batch_id: str
    ) -> Dict[str, Any]:
        """
        Transform raw row to standard format

        Demo row structure:
        - Brand: Product brand code
        - BrandName: Product name
        - Month: Month number (1-12)
        - Information: Additional info/description
        """
        # Extract fields
        brand = str(raw_row.get("Brand", "")).strip()
        brand_name = str(raw_row.get("BrandName", "")).strip()
        month = raw_row.get("Month")
        information = str(raw_row.get("Information", "")).strip()

        # Validate required fields
        if not brand or not brand_name:
            raise ValueError("Missing required fields: Brand or BrandName")

        # Parse month to create sale date
        # Default to current year if month is provided
        current_year = datetime.now().year
        try:
            month_int = int(month) if month else 1
            if not (1 <= month_int <= 12):
                month_int = 1
            sale_date = datetime(current_year, month_int, 1)
        except:
            sale_date = datetime.now()

        # Generate product_id and reseller_id
        product_id = str(uuid.uuid4())
        reseller_id = str(uuid.uuid4())

        # Create transformed record
        return {
            # Online Sale fields (D2C)
            "sale_id": str(uuid.uuid4()),
            "user_id": user_id,
            "batch_id": batch_id,
            "sale_date": sale_date.isoformat(),
            "product_ean": brand,
            "functional_name": brand_name,
            "product_id": product_id,
            "quantity": 1,  # Default quantity for demo
            "unit_price_eur": 100.0,  # Default price for demo
            "total_sales_eur": 100.0,
            "channel": "online",  # Demo is online sales
            "reseller_id": reseller_id,
            "reseller_name": "Demo Reseller",
            "country": "Demo",
            "metadata": {
                "information": information,
                "brand": brand,
                "month": month,
                "source": "demo_upload"
            },
            "created_at": datetime.now().isoformat()
        }

    @staticmethod
    def detect(file_path: str) -> bool:
        """
        Detect if file is Demo format

        Args:
            file_path: Path to file

        Returns:
            True if Demo format detected
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook[workbook.sheetnames[0]]

            # Get headers from first row
            headers = [str(cell.value).strip().lower() if cell.value else ""
                      for cell in sheet[1]]

            workbook.close()

            # Check for Demo-specific headers
            required_headers = {"brand", "brandname", "month", "information"}
            found_headers = {h.lower() for h in headers if h}

            return required_headers.issubset(found_headers)

        except Exception:
            return False

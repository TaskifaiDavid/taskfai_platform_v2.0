"""
Skins SA vendor data processor (South Africa - ZAR to EUR conversion)
"""

from typing import Dict, List, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class SkinsSAProcessor:
    """Process Skins SA Excel files with ZAR to EUR conversion"""

    TARGET_SHEET = "Sheet1"
    ZAR_TO_EUR_RATE = 0.05  # Approximate rate, should be configurable

    COLUMN_MAPPING = {
        "OrderDate": "order_date",
        "EAN": "product_ean",
        "Product": "functional_name",
        "Qty": "quantity",
        "Amount": "sales_zar",
        "Reseller": "reseller"
    }

    def __init__(self, exchange_rate: Optional[float] = None):
        """Initialize with optional custom exchange rate"""
        self.exchange_rate = exchange_rate or self.ZAR_TO_EUR_RATE

    def process(
        self,
        file_path: str,
        user_id: str,
        batch_id: str
    ) -> Dict[str, Any]:
        """
        Process Skins SA file with ZAR to EUR conversion

        Args:
            file_path: Path to Excel file
            user_id: User identifier
            batch_id: Batch identifier

        Returns:
            Processing result with statistics
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            if self.TARGET_SHEET not in workbook.sheetnames:
                sheet = workbook[workbook.sheetnames[0]]
            else:
                sheet = workbook[self.TARGET_SHEET]

            raw_rows = self._extract_rows(sheet)

            transformed_rows = []
            errors = []

            for row_num, raw_row in enumerate(raw_rows, start=2):
                try:
                    transformed = self._transform_row(raw_row, user_id, batch_id)
                    if transformed:
                        transformed_rows.append(transformed)
                except Exception as e:
                    errors.append({
                        "row_number": row_num,
                        "error": str(e),
                        "raw_data": raw_row
                    })

            workbook.close()

            return {
                "total_rows": len(raw_rows),
                "successful_rows": len(transformed_rows),
                "failed_rows": len(errors),
                "transformed_data": transformed_rows,
                "errors": errors
            }

        except Exception as e:
            raise Exception(f"Failed to process Skins SA file: {str(e)}")

    def _extract_rows(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """Extract rows from worksheet"""
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())

        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            row_dict = {}
            for idx, header in enumerate(headers):
                if idx < len(row):
                    row_dict[header] = row[idx]

            rows.append(row_dict)

        return rows

    def _transform_row(
        self,
        raw_row: Dict[str, Any],
        user_id: str,
        batch_id: str
    ) -> Optional[Dict[str, Any]]:
        """Transform raw row with ZAR to EUR conversion"""
        transformed = {
            "user_id": user_id,
            "batch_id": batch_id,
            "vendor": "skins_sa",
            "currency": "EUR"
        }

        for source_col, target_col in self.COLUMN_MAPPING.items():
            value = raw_row.get(source_col)

            if source_col in ["OrderDate", "EAN", "Qty", "Amount"] and not value:
                raise ValueError(f"Missing required field: {source_col}")

            if target_col == "product_ean":
                transformed[target_col] = self._validate_ean(value)
            elif target_col == "quantity":
                transformed[target_col] = self._to_int(value, "Qty")
            elif target_col == "sales_zar":
                zar_amount = self._to_float(value, "Amount")
                # Convert ZAR to EUR
                transformed["sales_eur"] = round(zar_amount * self.exchange_rate, 2)
            elif target_col == "order_date":
                # Extract month and year from order date
                date_obj = self._parse_date(value)
                transformed["month"] = date_obj.month
                transformed["year"] = date_obj.year
                transformed["order_date"] = date_obj.isoformat()
            elif target_col == "functional_name":
                transformed[target_col] = str(value)
            elif target_col == "reseller":
                transformed[target_col] = str(value) if value else "Skins SA"
            else:
                transformed[target_col] = str(value) if value is not None else None

        transformed["created_at"] = datetime.utcnow().isoformat()

        return transformed

    def _validate_ean(self, value: Any) -> str:
        """Validate EAN code"""
        if not value:
            raise ValueError("EAN cannot be empty")

        ean_str = str(value).strip()
        if '.' in ean_str:
            ean_str = ean_str.split('.')[0]

        if len(ean_str) != 13 or not ean_str.isdigit():
            raise ValueError(f"Invalid EAN format: {ean_str}")

        return ean_str

    def _parse_date(self, value: Any) -> datetime:
        """Parse date from various formats"""
        if isinstance(value, datetime):
            return value

        if isinstance(value, str):
            # Try common date formats
            for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]:
                try:
                    return datetime.strptime(value, fmt)
                except ValueError:
                    continue

        raise ValueError(f"Unable to parse date: {value}")

    def _to_int(self, value: Any, field_name: str) -> int:
        """Convert value to integer"""
        if value is None:
            raise ValueError(f"{field_name} cannot be None")
        try:
            return int(float(value))
        except (ValueError, TypeError):
            raise ValueError(f"Invalid integer for {field_name}: {value}")

    def _to_float(self, value: Any, field_name: str) -> float:
        """Convert value to float"""
        if value is None or value == "":
            raise ValueError(f"{field_name} cannot be None")
        try:
            return float(value)
        except (ValueError, TypeError):
            raise ValueError(f"Invalid float for {field_name}: {value}")

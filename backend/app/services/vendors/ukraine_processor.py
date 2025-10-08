"""
Ukraine vendor data processor (Ukraine - UAH to EUR conversion)
"""

from typing import Dict, List, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class UkraineProcessor:
    """Process Ukraine Excel files with UAH to EUR conversion"""

    TARGET_SHEET = "TDSheet"
    UAH_TO_EUR_RATE = 0.024  # Approximate rate, should be configurable

    COLUMN_MAPPING = {
        "EAN": "product_ean",
        "Product": "functional_name",
        "Quantity": "quantity",
        "Amount": "sales_uah",
        "Month": "month",
        "Year": "year"
    }

    def __init__(self, exchange_rate: Optional[float] = None):
        """Initialize with optional custom exchange rate"""
        self.exchange_rate = exchange_rate or self.UAH_TO_EUR_RATE

    def process(
        self,
        file_path: str,
        user_id: str,
        batch_id: str
    ) -> Dict[str, Any]:
        """
        Process Ukraine file with UAH to EUR conversion

        Args:
            file_path: Path to Excel file
            user_id: User identifier
            batch_id: Batch identifier

        Returns:
            Processing result with statistics
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            if self.TARGET_SHEET not in workbook.sheetnames:
                sheet = workbook[workbook.sheetnames[0]]
            else:
                sheet = workbook[self.TARGET_SHEET]

            raw_rows = self._extract_rows(sheet)

            transformed_rows = []
            errors = []

            for row_num, raw_row in enumerate(raw_rows, start=2):
                try:
                    transformed = self._transform_row(raw_row, user_id, batch_id)
                    if transformed:
                        transformed_rows.append(transformed)
                except Exception as e:
                    errors.append({
                        "row_number": row_num,
                        "error": str(e),
                        "raw_data": raw_row
                    })

            workbook.close()

            return {
                "total_rows": len(raw_rows),
                "successful_rows": len(transformed_rows),
                "failed_rows": len(errors),
                "transformed_data": transformed_rows,
                "errors": errors
            }

        except Exception as e:
            raise Exception(f"Failed to process Ukraine file: {str(e)}")

    def _extract_rows(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """Extract rows from worksheet"""
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())

        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            row_dict = {}
            for idx, header in enumerate(headers):
                if idx < len(row):
                    row_dict[header] = row[idx]

            rows.append(row_dict)

        return rows

    def _transform_row(
        self,
        raw_row: Dict[str, Any],
        user_id: str,
        batch_id: str
    ) -> Optional[Dict[str, Any]]:
        """Transform raw row with UAH to EUR conversion"""
        transformed = {
            "user_id": user_id,
            "batch_id": batch_id,
            "vendor": "ukraine",
            "currency": "EUR",
            "reseller": "Ukraine Distributors"
        }

        for source_col, target_col in self.COLUMN_MAPPING.items():
            value = raw_row.get(source_col)

            if source_col in ["EAN", "Product", "Quantity"] and not value:
                raise ValueError(f"Missing required field: {source_col}")

            if target_col == "product_ean":
                transformed[target_col] = self._validate_ean(value)
            elif target_col == "quantity":
                transformed[target_col] = self._to_int(value, "Quantity")
            elif target_col == "sales_uah":
                uah_amount = self._to_float(value, "Amount")
                # Convert UAH to EUR
                transformed["sales_eur"] = round(uah_amount * self.exchange_rate, 2)
            elif target_col == "month":
                transformed[target_col] = self._validate_month(value) if value else datetime.utcnow().month
            elif target_col == "year":
                transformed[target_col] = self._validate_year(value) if value else datetime.utcnow().year
            elif target_col == "functional_name":
                transformed[target_col] = str(value)
            else:
                transformed[target_col] = str(value) if value is not None else None

        transformed["created_at"] = datetime.utcnow().isoformat()

        return transformed

    def _validate_ean(self, value: Any) -> str:
        """Validate EAN code"""
        if not value:
            raise ValueError("EAN cannot be empty")

        ean_str = str(value).strip()
        if '.' in ean_str:
            ean_str = ean_str.split('.')[0]

        if len(ean_str) != 13 or not ean_str.isdigit():
            raise ValueError(f"Invalid EAN format: {ean_str}")

        return ean_str

    def _validate_month(self, value: Any) -> int:
        """Validate month value"""
        month = self._to_int(value, "Month")
        if month < 1 or month > 12:
            raise ValueError(f"Invalid month: {month}")
        return month

    def _validate_year(self, value: Any) -> int:
        """Validate year value"""
        year = self._to_int(value, "Year")
        if year < 2000 or year > 2100:
            raise ValueError(f"Invalid year: {year}")
        return year

    def _to_int(self, value: Any, field_name: str) -> int:
        """Convert value to integer"""
        if value is None:
            raise ValueError(f"{field_name} cannot be None")
        try:
            return int(float(value))
        except (ValueError, TypeError):
            raise ValueError(f"Invalid integer for {field_name}: {value}")

    def _to_float(self, value: Any, field_name: str) -> float:
        """Convert value to float"""
        if value is None or value == "":
            return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            raise ValueError(f"Invalid float for {field_name}: {value}")

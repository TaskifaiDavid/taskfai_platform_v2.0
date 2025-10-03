"""
Boxnox vendor data processor
"""

from typing import Dict, List, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class BoxnoxProcessor:
    """Process Boxnox Excel files"""

    TARGET_SHEET = "Sell Out by EAN"

    COLUMN_MAPPING = {
        "Product EAN": "product_ean",
        "Functional Name": "functional_name",
        "Sold Qty": "quantity",
        "Sales Amount (EUR)": "sales_eur",
        "Reseller": "reseller",
        "Month": "month",
        "Year": "year"
    }

    def process(
        self,
        file_path: str,
        user_id: str,
        batch_id: str
    ) -> Dict[str, Any]:
        """
        Process Boxnox file

        Args:
            file_path: Path to Excel file
            user_id: User identifier
            batch_id: Batch identifier

        Returns:
            Processing result with statistics
        """
        try:
            # Load workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            # Check if target sheet exists
            if self.TARGET_SHEET not in workbook.sheetnames:
                raise ValueError(f"Sheet '{self.TARGET_SHEET}' not found in workbook")

            sheet = workbook[self.TARGET_SHEET]

            # Extract data
            raw_rows = self._extract_rows(sheet)

            # Transform data
            transformed_rows = []
            errors = []

            for row_num, raw_row in enumerate(raw_rows, start=2):  # Start from 2 (after header)
                try:
                    transformed = self._transform_row(raw_row, user_id, batch_id)
                    if transformed:
                        transformed_rows.append(transformed)
                except Exception as e:
                    errors.append({
                        "row_number": row_num,
                        "error": str(e),
                        "raw_data": raw_row
                    })

            workbook.close()

            return {
                "total_rows": len(raw_rows),
                "successful_rows": len(transformed_rows),
                "failed_rows": len(errors),
                "transformed_data": transformed_rows,
                "errors": errors
            }

        except Exception as e:
            raise Exception(f"Failed to process Boxnox file: {str(e)}")

    def _extract_rows(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """
        Extract rows from worksheet

        Args:
            sheet: Excel worksheet

        Returns:
            List of row dictionaries
        """
        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())

        # Extract data rows
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Skip empty rows
            if not any(row):
                continue

            row_dict = {}
            for idx, header in enumerate(headers):
                if idx < len(row):
                    row_dict[header] = row[idx]

            rows.append(row_dict)

        return rows

    def _transform_row(
        self,
        raw_row: Dict[str, Any],
        user_id: str,
        batch_id: str
    ) -> Optional[Dict[str, Any]]:
        """
        Transform raw row to standardized format

        Args:
            raw_row: Raw row data
            user_id: User identifier
            batch_id: Batch identifier

        Returns:
            Transformed row or None if invalid
        """
        # Map columns
        transformed = {
            "user_id": user_id,
            "batch_id": batch_id,
            "vendor": "boxnox"
        }

        for source_col, target_col in self.COLUMN_MAPPING.items():
            value = raw_row.get(source_col)

            # Handle missing required fields
            if source_col in ["Product EAN", "Sold Qty", "Month", "Year"] and not value:
                raise ValueError(f"Missing required field: {source_col}")

            # Type conversions and validations
            if target_col == "product_ean":
                transformed[target_col] = self._validate_ean(value)
            elif target_col == "quantity":
                transformed[target_col] = self._to_int(value, "Sold Qty")
            elif target_col == "sales_eur":
                transformed[target_col] = self._to_float(value, "Sales Amount (EUR)")
            elif target_col == "month":
                transformed[target_col] = self._validate_month(value)
            elif target_col == "year":
                transformed[target_col] = self._validate_year(value)
            else:
                transformed[target_col] = str(value) if value is not None else None

        # Add metadata
        transformed["created_at"] = datetime.utcnow().isoformat()

        return transformed

    def _validate_ean(self, value: Any) -> str:
        """Validate EAN code"""
        if not value:
            raise ValueError("EAN cannot be empty")

        ean_str = str(value).strip()

        # Remove any decimal points (in case Excel formatted as number)
        if '.' in ean_str:
            ean_str = ean_str.split('.')[0]

        # Check length (13 digits for EAN-13)
        if len(ean_str) != 13 or not ean_str.isdigit():
            raise ValueError(f"Invalid EAN format: {ean_str}")

        return ean_str

    def _validate_month(self, value: Any) -> int:
        """Validate month value"""
        month = self._to_int(value, "Month")

        if month < 1 or month > 12:
            raise ValueError(f"Invalid month: {month}")

        return month

    def _validate_year(self, value: Any) -> int:
        """Validate year value"""
        year = self._to_int(value, "Year")

        if year < 2000 or year > 2100:
            raise ValueError(f"Invalid year: {year}")

        return year

    def _to_int(self, value: Any, field_name: str) -> int:
        """Convert value to integer"""
        if value is None:
            raise ValueError(f"{field_name} cannot be None")

        try:
            return int(float(value))
        except (ValueError, TypeError):
            raise ValueError(f"Invalid integer for {field_name}: {value}")

    def _to_float(self, value: Any, field_name: str) -> Optional[float]:
        """Convert value to float"""
        if value is None or value == "":
            return None

        try:
            return float(value)
        except (ValueError, TypeError):
            raise ValueError(f"Invalid float for {field_name}: {value}")

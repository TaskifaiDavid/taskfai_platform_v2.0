"""
Galilu Processor (BIBBI Version)

Handles Galilu Poland reseller data with special cases:
1. No EAN codes: Uses product names that need mapping to EANs
2. Multi-sheet stores: Each Excel sheet represents a different store
3. Price not in file: Needs to be fetched from products table
4. Currency: PLN → EUR conversion

Based on: backend/BIBBI/Resellers/resellers_info.md
"""

from typing import List, Dict, Any, Optional
from datetime import datetime
import openpyxl

from .base import BibbiBseProcessor
from app.core.bibbi import BibbιDB
from app.services.bibbi.product_mapping_service import BibbιProductMappingService


class GaliluProcessor(BibbiBseProcessor):
    """Process Galilu Excel files with product name matching"""

    VENDOR_NAME = "galilu"
    CURRENCY = "PLN"
    PLN_TO_EUR_RATE = 0.23

    # Column mapping from Galilu format to internal names
    COLUMN_MAPPING = {
        "Product": "product_name",
        "Product Name": "product_name",
        "Quantity": "quantity",
        "Qty": "quantity",
        "Month": "month",
        "Year": "year",
    }

    def __init__(self, reseller_id: str, bibbi_db: Optional[BibbιDB] = None):
        """
        Initialize Galilu processor

        Args:
            reseller_id: UUID of Galilu reseller
            bibbi_db: BIBBI database client for product mapping lookups
        """
        super().__init__(reseller_id)
        self.bibbi_db = bibbi_db
        # Initialize product mapping service for product name → EAN mapping
        self.product_mapping_service = BibbιProductMappingService(bibbi_db) if bibbi_db else None

    def get_vendor_name(self) -> str:
        return self.VENDOR_NAME

    def get_currency(self) -> str:
        return self.CURRENCY

    def extract_stores(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Extract Galilu store information

        CRITICAL: Each Excel sheet represents a different store
        Sheet name = store name

        Returns:
            List of store dictionaries (one per sheet)
        """
        stores = []

        try:
            workbook = self._load_workbook(file_path)

            for sheet_name in workbook.sheetnames:
                # Skip system/hidden sheets
                if sheet_name.startswith('_') or sheet_name.lower() in ['info', 'metadata']:
                    continue

                # Normalize store name
                store_identifier = sheet_name.strip().lower().replace(' ', '_')

                stores.append({
                    "store_identifier": store_identifier,
                    "store_name": f"Galilu {sheet_name}",
                    "store_type": "physical",  # Galilu stores are physical retail
                    "reseller_id": self.reseller_id,
                    "country": "Poland"
                })

            workbook.close()

            if not stores:
                # Fallback: If no valid sheets, create single store
                stores = [{
                    "store_identifier": "galilu_main",
                    "store_name": "Galilu Main Store",
                    "store_type": "physical",
                    "reseller_id": self.reseller_id,
                    "country": "Poland"
                }]

        except Exception as e:
            print(f"[Galilu] Error extracting stores: {e}")
            # Fallback
            stores = [{
                "store_identifier": "galilu_main",
                "store_name": "Galilu Main Store",
                "store_type": "physical",
                "reseller_id": self.reseller_id,
                "country": "Poland"
            }]

        return stores

    def extract_rows(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Extract rows from Galilu Excel file

        CRITICAL: Must process ALL sheets (each = different store)
        Each row gets tagged with sheet_name for store mapping
        """
        all_rows = []

        workbook = self._load_workbook(file_path)

        for sheet_name in workbook.sheetnames:
            # Skip system sheets
            if sheet_name.startswith('_') or sheet_name.lower() in ['info', 'metadata']:
                continue

            sheet = workbook[sheet_name]
            headers = self._get_sheet_headers(sheet)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                row_dict = {"_sheet_name": sheet_name}  # Store sheet name for later
                for idx, header in enumerate(headers):
                    if idx < len(row):
                        row_dict[header] = row[idx]

                all_rows.append(row_dict)

        workbook.close()
        return all_rows

    def transform_row(
        self,
        raw_row: Dict[str, Any],
        batch_id: str
    ) -> Optional[Dict[str, Any]]:
        """
        Transform Galilu row to sales_unified schema

        CRITICAL:
        1. Product name → EAN mapping required (query product_reseller_mappings)
        2. Sheet name → store identifier
        3. Price from products table (list_price)
        4. PLN → EUR conversion
        """
        # Start with base row
        transformed = self._create_base_row(batch_id)

        # Extract product name (NO EAN in file)
        product_name = raw_row.get("Product") or raw_row.get("Product Name")
        if not product_name:
            raise ValueError("Missing Product name")

        product_name = str(product_name).strip()
        transformed["product_name_raw"] = product_name

        # Map product name → EAN via product_reseller_mappings
        ean = self._match_product_name_to_ean(product_name)
        if not ean:
            raise ValueError(f"Product not mapped: {product_name}")

        transformed["product_id"] = ean

        # Extract quantity
        qty_value = raw_row.get("Quantity") or raw_row.get("Qty")
        if qty_value is None:
            raise ValueError("Missing Quantity")

        try:
            quantity = self._to_int(qty_value, "Quantity")
            if quantity <= 0:
                raise ValueError(f"Invalid quantity: {quantity}")

            transformed["quantity"] = quantity
            transformed["is_return"] = False

        except ValueError as e:
            raise ValueError(f"Invalid quantity: {e}")

        # Get price from products table (Galilu doesn't report sales amount)
        list_price = self._get_product_list_price(ean)
        if list_price is None:
            raise ValueError(f"Product price not found for EAN: {ean}")

        # Calculate sales amount
        sales_pln = list_price * quantity
        transformed["sales_local_currency"] = sales_pln
        transformed["sales_eur"] = self._convert_currency(sales_pln, "PLN")

        # Extract date information
        month_value = raw_row.get("Month")
        year_value = raw_row.get("Year")

        if month_value and year_value:
            try:
                month = self._to_int(month_value, "Month")
                year = self._to_int(year_value, "Year")

                if month < 1 or month > 12:
                    raise ValueError(f"Invalid month: {month}")

                if year < 2000 or year > 2100:
                    raise ValueError(f"Invalid year: {year}")

                # Create sale_date (use first day of month)
                transformed["sale_date"] = datetime(year, month, 1).date().isoformat()
                transformed["year"] = year
                transformed["month"] = month
                transformed["quarter"] = self._calculate_quarter(month)

            except ValueError as e:
                raise ValueError(f"Invalid date: {e}")
        else:
            # Use current date if not provided
            now = datetime.utcnow()
            transformed["sale_date"] = now.date().isoformat()
            transformed["year"] = now.year
            transformed["month"] = now.month
            transformed["quarter"] = self._calculate_quarter(now.month)

        # Extract store from sheet name
        sheet_name = raw_row.get("_sheet_name", "Sheet1")
        store_identifier = sheet_name.strip().lower().replace(' ', '_')
        transformed["store_identifier"] = store_identifier

        return transformed

    def _match_product_name_to_ean(self, product_name: str) -> Optional[str]:
        """
        Match Galilu product name to EAN via product mapping service

        Uses BibbιProductMappingService for:
        - Exact matching
        - Fuzzy matching (85% similarity threshold)
        - Caching for performance

        Args:
            product_name: Galilu product name

        Returns:
            EAN (product_id) or None if not found
        """
        if not self.product_mapping_service:
            print(f"[Galilu] Warning: No product mapping service available")
            return None

        try:
            # Use product mapping service with fuzzy matching enabled
            ean = self.product_mapping_service.get_ean_by_product_code(
                reseller_id=self.reseller_id,
                product_code=product_name,
                use_fuzzy_match=True  # Enable fuzzy matching for variations
            )

            if ean:
                print(f"[Galilu] Mapped product: '{product_name}' → {ean}")
            else:
                print(f"[Galilu] Product not mapped: '{product_name}'")

            return ean

        except Exception as e:
            print(f"[Galilu] Error matching product: {e}")
            return None

    def _get_product_list_price(self, ean: str) -> Optional[float]:
        """
        Get product list price from products table

        Query: products
        WHERE ean = ean

        Args:
            ean: Product EAN

        Returns:
            List price or None if not found
        """
        if not self.bibbi_db:
            print(f"[Galilu] Warning: No database client for price lookup")
            return None

        try:
            # Query products table
            # NOTE: Use raw client to bypass tenant filter (products table has no tenant_id)
            result = self.bibbi_db.client.table("products")\
                .select("list_price")\
                .eq("ean", ean)\
                .execute()

            if result.data and len(result.data) > 0:
                list_price = result.data[0].get("list_price")
                if list_price:
                    return float(list_price)

            return None

        except Exception as e:
            print(f"[Galilu] Error getting product price: {e}")
            return None


def get_galilu_processor(reseller_id: str, bibbi_db: BibbιDB) -> GaliluProcessor:
    """
    Factory function to create Galilu processor

    Args:
        reseller_id: UUID of Galilu reseller from resellers table
        bibbi_db: BIBBI database client for product/price lookups

    Returns:
        GaliluProcessor instance
    """
    return GaliluProcessor(reseller_id, bibbi_db)

"""
BIBBI Base Processor

Abstract base class for all BIBBI reseller data processors.

All processors must:
1. Extract rows from Excel files
2. Transform rows to sales_unified schema
3. Extract store information
4. Validate transformed data
5. Handle currency conversion
"""

from abc import ABC, abstractmethod
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
from decimal import Decimal, InvalidOperation
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.core.bibbi import BIBBI_TENANT_ID
from app.utils.validation import validate_ean, to_int, to_float
from app.utils.excel import (
    extract_rows_from_sheet,
    get_sheet_headers,
    safe_load_workbook
)


class ProcessingResult:
    """Result of file processing"""

    def __init__(
        self,
        vendor: str,
        total_rows: int,
        successful_rows: int,
        failed_rows: int,
        transformed_data: List[Dict[str, Any]],
        stores: List[Dict[str, Any]],
        errors: List[Dict[str, Any]]
    ):
        self.vendor = vendor
        self.total_rows = total_rows
        self.successful_rows = successful_rows
        self.failed_rows = failed_rows
        self.transformed_data = transformed_data
        self.stores = stores
        self.errors = errors

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for JSON serialization"""
        return {
            "vendor": self.vendor,
            "total_rows": self.total_rows,
            "successful_rows": self.successful_rows,
            "failed_rows": self.failed_rows,
            "transformed_data": self.transformed_data,
            "stores": self.stores,
            "errors": self.errors
        }


class BibbiBseProcessor(ABC):
    """
    Abstract base class for BIBBI vendor processors

    Each vendor processor must implement:
    - extract_rows(): Parse Excel file into raw dictionaries
    - transform_row(): Convert raw row to sales_unified schema
    - extract_stores(): Identify and extract store records
    - get_vendor_name(): Return vendor identifier
    """

    # Currency conversion rates (approximate - should be configurable)
    CURRENCY_RATES = {
        "EUR": 1.0,  # Base currency
        "GBP": 1.17,  # GBP to EUR
        "USD": 0.92,  # USD to EUR
        "PLN": 0.23,  # Polish Zloty to EUR
        "ZAR": 0.05,  # South African Rand to EUR
    }

    def __init__(self, reseller_id: str):
        """
        Initialize processor

        Args:
            reseller_id: UUID of reseller from resellers table
        """
        self.reseller_id = reseller_id
        self.tenant_id = BIBBI_TENANT_ID
        # Cache for reseller details to avoid repeated queries
        self._reseller_cache: Optional[Dict[str, Any]] = None

    @abstractmethod
    def get_vendor_name(self) -> str:
        """Return vendor identifier (e.g., 'liberty', 'galilu')"""
        pass

    @abstractmethod
    def get_currency(self) -> str:
        """Return vendor's reporting currency (EUR, GBP, PLN, ZAR)"""
        pass

    @abstractmethod
    def extract_rows(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Extract raw rows from Excel file

        Args:
            file_path: Path to Excel file

        Returns:
            List of raw row dictionaries with vendor-specific column names
        """
        pass

    @abstractmethod
    def transform_row(
        self,
        raw_row: Dict[str, Any],
        batch_id: str
    ) -> Optional[Dict[str, Any]]:
        """
        Transform raw row to sales_unified schema

        Args:
            raw_row: Raw row from extract_rows()
            batch_id: Batch identifier for this upload

        Returns:
            Transformed row dict or None if row should be skipped

        Raises:
            ValueError: If row validation fails
        """
        pass

    @abstractmethod
    def extract_stores(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Extract store information from file

        Args:
            file_path: Path to Excel file

        Returns:
            List of store dictionaries:
            {
                "store_identifier": str,  # Vendor's store code/name
                "store_name": str,        # Display name
                "store_type": str,        # "physical" or "online"
                "reseller_id": str        # UUID
            }
        """
        pass

    def process(
        self,
        file_path: str,
        batch_id: str
    ) -> ProcessingResult:
        """
        Main processing pipeline

        1. Extract stores
        2. Extract raw rows
        3. Transform each row
        4. Collect errors
        5. Return results

        Args:
            file_path: Path to Excel file
            batch_id: Batch identifier

        Returns:
            ProcessingResult with transformed data, stores, and errors
        """
        vendor = self.get_vendor_name()
        print(f"[{vendor}] Starting processing: {file_path}")

        # Extract stores first (needed for store_id mapping)
        try:
            stores = self.extract_stores(file_path)
            print(f"[{vendor}] Extracted {len(stores)} stores")
        except Exception as e:
            print(f"[{vendor}] Error extracting stores: {e}")
            stores = []

        # Extract raw rows
        try:
            raw_rows = self.extract_rows(file_path)
            total_rows = len(raw_rows)
            print(f"[{vendor}] Extracted {total_rows} rows")
        except Exception as e:
            print(f"[{vendor}] Error extracting rows: {e}")
            return ProcessingResult(
                vendor=vendor,
                total_rows=0,
                successful_rows=0,
                failed_rows=0,
                transformed_data=[],
                stores=stores,
                errors=[{"error": f"Failed to extract rows: {str(e)}"}]
            )

        # Transform rows
        transformed_data = []
        errors = []

        for row_num, raw_row in enumerate(raw_rows, start=2):  # Start at 2 (Excel row numbers, skip header)
            try:
                transformed = self.transform_row(raw_row, batch_id)
                if transformed:
                    # Inject tenant_id
                    transformed["tenant_id"] = self.tenant_id
                    transformed_data.append(transformed)
            except Exception as e:
                errors.append({
                    "row_number": row_num,
                    "error": str(e),
                    "raw_data": raw_row
                })

        successful_rows = len(transformed_data)
        failed_rows = len(errors)

        print(f"[{vendor}] Processing complete: {successful_rows} success, {failed_rows} failed")

        return ProcessingResult(
            vendor=vendor,
            total_rows=total_rows,
            successful_rows=successful_rows,
            failed_rows=failed_rows,
            transformed_data=transformed_data,
            stores=stores,
            errors=errors
        )

    # Utility methods for common operations
    # NOTE: Common utilities now imported from app.utils.validation and app.utils.excel

    def _load_workbook(self, file_path: str, read_only: bool = True):
        """
        Load Excel workbook safely

        Uses shared utility: app.utils.excel.safe_load_workbook
        """
        return safe_load_workbook(file_path, data_only=True, read_only=read_only)

    def _get_sheet_headers(self, sheet: Worksheet) -> List[str]:
        """
        Extract column headers from first row

        Uses shared utility: app.utils.excel.get_sheet_headers
        """
        return get_sheet_headers(sheet, header_row=1)

    def _validate_ean(self, value: Any, required: bool = True) -> Optional[str]:
        """
        Validate and normalize EAN code

        Uses shared utility: app.utils.validation.validate_ean

        Args:
            value: EAN value from Excel
            required: If False, returns None for empty values

        Returns:
            Normalized 13-digit EAN string or None if not required

        Raises:
            ValueError: If EAN is invalid and required=True
        """
        return validate_ean(value, required=required, strict=True)

    def _to_int(self, value: Any, field_name: str) -> int:
        """
        Convert value to integer safely

        Uses shared utility: app.utils.validation.to_int
        Extended with accounting notation support: "(123)" = -123

        Args:
            value: Value to convert
            field_name: Field name for error messages

        Returns:
            Integer value

        Raises:
            ValueError: If conversion fails
        """
        # Handle accounting notation: "(123)" means negative
        if isinstance(value, str):
            value = value.strip()
            if value.startswith('(') and value.endswith(')'):
                value = '-' + value[1:-1]

        # Use shared utility for standard conversion
        return to_int(value, field_name)

    def _to_float(self, value: Any, field_name: str) -> float:
        """
        Convert value to float safely

        Uses shared utility: app.utils.validation.to_float
        Extended with accounting notation support: "(123.45)" = -123.45

        Args:
            value: Value to convert
            field_name: Field name for error messages

        Returns:
            Float value (defaults to 0.0 for None/empty)

        Raises:
            ValueError: If conversion fails
        """
        # Handle accounting notation: "(123.45)" means negative
        if isinstance(value, str):
            value = value.strip()
            if value.startswith('(') and value.endswith(')'):
                value = '-' + value[1:-1]

        # Use shared utility with allow_none=True, default=0.0
        return to_float(value, field_name, allow_none=True, default=0.0)

    def _to_decimal(self, value: Any, field_name: str) -> Decimal:
        """
        Convert value to Decimal safely

        Args:
            value: Value to convert
            field_name: Field name for error messages

        Returns:
            Decimal value

        Raises:
            ValueError: If conversion fails
        """
        try:
            return Decimal(str(self._to_float(value, field_name)))
        except (ValueError, InvalidOperation):
            raise ValueError(f"Invalid decimal for {field_name}: {value}")

    def _convert_currency(self, amount: float, from_currency: str) -> float:
        """
        Convert amount to EUR

        Args:
            amount: Amount in source currency
            from_currency: Source currency code (EUR, GBP, PLN, ZAR)

        Returns:
            Amount in EUR
        """
        if from_currency == "EUR":
            return amount

        rate = self.CURRENCY_RATES.get(from_currency)
        if rate is None:
            raise ValueError(f"Unknown currency: {from_currency}")

        return round(amount * rate, 2)

    def _validate_date(self, value: Any) -> datetime:
        """
        Validate and parse date value

        Args:
            value: Date value (datetime object or string)

        Returns:
            datetime object

        Raises:
            ValueError: If date is invalid
        """
        if isinstance(value, datetime):
            return value

        if isinstance(value, str):
            # Try common date formats
            for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]:
                try:
                    return datetime.strptime(value, fmt)
                except ValueError:
                    continue

            raise ValueError(f"Invalid date format: {value}")

        raise ValueError(f"Invalid date type: {type(value)}")

    def _calculate_quarter(self, month: int) -> int:
        """Calculate quarter from month (1-4)"""
        return (month - 1) // 3 + 1

    def _get_reseller_sales_channel(self) -> Optional[str]:
        """
        Fetch sales_channel from resellers table (with caching)

        Returns:
            sales_channel value ("B2B", "B2C", "B2B2C", etc.) or None
        """
        # Check cache first
        if self._reseller_cache is not None:
            return self._reseller_cache.get("sales_channel")

        try:
            # Import here to avoid circular dependency
            from app.core.bibbi import get_bibbi_db

            bibbi_db = get_bibbi_db()

            # NOTE: Use raw client to bypass tenant filter (resellers table has no tenant_id)
            result = bibbi_db.client.table("resellers")\
                .select("sales_channel, reseller")\
                .eq("id", self.reseller_id)\
                .execute()

            if result.data and len(result.data) > 0:
                # Cache for subsequent calls
                self._reseller_cache = result.data[0]
                return self._reseller_cache.get("sales_channel")
            else:
                print(f"[BibbiProcessor] Reseller {self.reseller_id} not found in resellers table")
                return None

        except Exception as e:
            print(f"[BibbiProcessor] Error fetching reseller details: {e}")
            return None

    def _create_base_row(self, batch_id: str) -> Dict[str, Any]:
        """
        Create base row with common fields

        Returns dict with:
        - tenant_id
        - reseller_id
        - batch_id
        - vendor_name
        - currency
        - sales_channel (from resellers table, can be overridden by child processors)
        - created_at

        NOTE: Child processors (like LibertyProcessor) can override sales_channel
        if their business logic requires a different semantic (e.g., distribution channel
        vs. business model).
        """
        base_row = {
            "tenant_id": self.tenant_id,
            "reseller_id": self.reseller_id,
            "batch_id": batch_id,
            "vendor_name": self.get_vendor_name(),
            "currency": self.get_currency(),
            "created_at": datetime.utcnow().isoformat()
        }

        # Fetch and add sales_channel from resellers table (default business model)
        # Child processors may override this for vendor-specific semantics
        sales_channel = self._get_reseller_sales_channel()
        if sales_channel:
            base_row["sales_channel"] = sales_channel
        else:
            # Default to B2B for reseller data if not explicitly set
            # Reseller uploads are typically B2B by nature
            print(f"[BibbiBseProcessor] WARNING: Reseller {self.reseller_id} missing sales_channel, using 'B2B' as default")
            base_row["sales_channel"] = "B2B"

        return base_row

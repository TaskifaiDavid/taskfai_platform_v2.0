"""
Liberty Processor (BIBBI Version)

Handles Liberty UK reseller data with special cases:
1. Duplicate rows: Liberty uses 2 rows per product with same amount
2. Returns: Parentheses indicate returns (123) = -123
3. Store identification: "Flagship" = physical, "Internet" = online
4. Currency: GBP → EUR conversion
5. Date extraction: Parses sale date from filename (DD_MM_YYYY.xlsx)

Based on: backend/BIBBI/Resellers/resellers_info.md
"""

from typing import List, Dict, Any, Optional
from datetime import datetime
import openpyxl
import re
import hashlib
from supabase import Client

from .base import BibbiBseProcessor


class LibertyProcessor(BibbiBseProcessor):
    """Process Liberty Excel files with GBP to EUR conversion"""

    VENDOR_NAME = "liberty"
    CURRENCY = "GBP"
    GBP_TO_EUR_RATE = 1.17

    # Column mapping from Liberty format to internal names
    # NOTE: Liberty has 3-row headers (rows 1-3), actual data starts at row 4
    COLUMN_MAPPING = {
        "Item ID | Colour": "product_ean",  # Contains EAN like "000834429 | 98-NO COLOUR"
        "Item": "functional_name",  # Product name
        "Sales Qty Un": "quantity",  # Quantity sold (may be in store-specific columns)
        "Sales Inc VAT £ ": "sales_gbp",  # Sales amount in GBP
        "Sales Inc VAT £": "sales_gbp",  # Alternative without trailing space
    }

    # Store-specific column pattern
    # Format: "Flagship" -> "Sales Qty Un", "Internet" -> "Sales Qty Un"
    # We need to detect which columns belong to which store dynamically

    def __init__(self, reseller_id: str, supabase_client: Client):
        """Initialize Liberty processor with cached services and product lookup table"""
        super().__init__(reseller_id)

        # Store Supabase client for product lookups
        self.supabase = supabase_client

        # Pre-load Liberty products into memory for fast lookups
        # Query products table for all Liberty products
        # Format: {liberty_name: {ean, functional_name}}
        self.liberty_products = {}
        try:
            result = self.supabase.table("products")\
                .select("liberty_name, ean, functional_name")\
                .not_.is_("liberty_name", "null")\
                .execute()

            for product in result.data:
                liberty_name = product.get('liberty_name')
                if liberty_name:
                    self.liberty_products[liberty_name] = {
                        'ean': product['ean'],
                        'functional_name': product['functional_name']
                    }

            print(f"[Liberty] Pre-loaded {len(self.liberty_products)} products from products table")
        except Exception as e:
            print(f"[Liberty] WARNING: Failed to pre-load products: {e}")
            self.liberty_products = {}

        # Track unmatched Liberty names for reporting
        self.unmatched_liberty_names = []

    def get_vendor_name(self) -> str:
        return self.VENDOR_NAME

    def get_currency(self) -> str:
        return self.CURRENCY

    def extract_stores(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Extract Liberty store information

        Liberty typically has:
        - "Flagship" = physical store (London flagship location)
        - "online" / "Online" = e-commerce
        - Sometimes "Store 1", "Store 2", etc.

        Returns:
            List of store dictionaries
        """
        stores = []

        try:
            workbook = self._load_workbook(file_path)
            sheet = workbook[workbook.sheetnames[0]]
            headers = self._get_sheet_headers(sheet)

            # Find store column
            store_col_idx = None
            for idx, header in enumerate(headers):
                if header in ["Store", "Location", "Shop"]:
                    store_col_idx = idx
                    break

            if store_col_idx is None:
                # No store column - assume single store
                # Liberty minimum: Flagship + internet
                stores = [
                    {
                        "store_identifier": "flagship",
                        "store_name": "Liberty Flagship",
                        "store_type": "physical",
                        "reseller_id": self.reseller_id,
                        "city": "London",
                        "country": "UK"
                    },
                    {
                        "store_identifier": "internet",
                        "store_name": "Liberty Online",
                        "store_type": "online",
                        "reseller_id": self.reseller_id,
                        "country": "UK"
                    }
                ]
            else:
                # Extract unique store identifiers
                seen_stores = set()
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue

                    store_value = row[store_col_idx] if store_col_idx < len(row) else None
                    if store_value:
                        store_str = str(store_value).strip().lower()
                        if store_str and store_str not in seen_stores:
                            seen_stores.add(store_str)

                            # Determine store type
                            is_online = any(keyword in store_str for keyword in ["online", "web", "e-commerce", "ecom"])
                            store_type = "online" if is_online else "physical"

                            stores.append({
                                "store_identifier": store_str,
                                "store_name": f"Liberty {store_value.strip()}",
                                "store_type": store_type,
                                "reseller_id": self.reseller_id,
                                "city": "London" if store_type == "physical" else None,
                                "country": "UK"
                            })

            workbook.close()

        except Exception as e:
            print(f"[Liberty] Error extracting stores: {e}")
            # Fallback to default stores
            stores = [
                {
                    "store_identifier": "flagship",
                    "store_name": "Liberty Flagship",
                    "store_type": "physical",
                    "reseller_id": self.reseller_id,
                    "city": "London",
                    "country": "UK"
                },
                {
                    "store_identifier": "internet",
                    "store_name": "Liberty Online",
                    "store_type": "online",
                    "reseller_id": self.reseller_id,
                    "country": "UK"
                }
            ]

        return stores

    def _parse_store_columns(self, sheet) -> Dict[str, Dict[str, int]]:
        """
        Parse Liberty's multi-store column structure from rows 1-3

        Row 1: Store names (Flagship, Internet, etc.)
        Row 2: Date range labels (Actual, YTD, etc.) - CRITICAL for filtering
        Row 3: Column headers (Sales Qty Un, Sales Inc VAT £, etc.)

        Returns:
            Dict mapping store identifier to column ranges for "Actual" columns only
            Example: {
                "flagship": {"qty_col": 12, "sales_col": 13},
                "internet": {"qty_col": 14, "sales_col": 15}
            }
        """
        store_columns = {}

        # Row 1 has store names like "Flagship", "Internet"
        # Row 2 has date range labels like "Actual", "YTD" - we ONLY want "Actual"
        # Row 3 has column headers like "Sales Qty Un", "Sales Inc VAT £ "
        row1 = list(sheet[1])
        row2 = list(sheet[2])
        row3 = list(sheet[3])

        current_store = None
        current_row2_label = ""  # Track most recent Row 2 value (handles merged cells)

        for idx, cell in enumerate(row1):
            if cell.value and str(cell.value).strip():
                store_name = str(cell.value).strip()

                # Skip non-store headers
                if store_name in ['Retail Group', 'Brand', 'Colour Phase', 'Product Group', 'Item ID | Colour', 'Item', 'All Warehouse', '']:
                    continue

                # Skip "All Sales Channels" - we want individual stores only
                if store_name.lower() in ['all sales channels', 'total']:
                    continue

                # Normalize store name to identifier
                store_id = store_name.lower().replace(' ', '_')
                if store_id not in store_columns:
                    store_columns[store_id] = {}
                    current_store = store_id

            # Update current Row 2 label when we find a value
            # This handles merged cells: when Row 2[idx] is None, we use the last seen label
            if idx < len(row2) and row2[idx].value:
                current_row2_label = str(row2[idx].value).strip().lower()

            # Find quantity and sales columns under this store
            # CRITICAL FIX: Only map columns in "Actual" sections
            # Use current_row2_label instead of checking current cell (handles merged cells)
            if current_store and idx < len(row3) and row3[idx].value:
                # Skip this column if we're not in an "Actual" section
                if 'actual' not in current_row2_label:
                    continue

                # Now check Row 3 header and map columns
                header = str(row3[idx].value).strip()

                if 'qty' in header.lower() or 'quantity' in header.lower():
                    # Only set if not already set (take first "Actual" occurrence)
                    if 'qty_col' not in store_columns[current_store]:
                        store_columns[current_store]['qty_col'] = idx
                elif 'sales' in header.lower() and '£' in header:
                    # Only set if not already set (take first "Actual" occurrence)
                    if 'sales_col' not in store_columns[current_store]:
                        store_columns[current_store]['sales_col'] = idx

        print(f"[Liberty] Parsed store columns: {store_columns}")
        return store_columns

    def extract_rows(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Extract rows from Liberty Excel file

        Liberty has complex structure:
        - Rows 1-2: Multi-level store headers (Flagship, Internet, etc.)
        - Row 3: Column headers (Item ID | Colour, Item, Sales Qty Un, etc.)
        - Row 4+: Data rows where each product uses TWO rows:
          - Row 1: Product info (brand, EAN, name) - columns A-L
          - Row 2: Sales data (quantities/amounts by store) - columns M onwards

        NEW: Creates MULTIPLE records per product - one for each store with data
        """
        workbook = self._load_workbook(file_path, read_only=False)
        sheet = workbook[workbook.sheetnames[0]]

        # Parse store column structure
        store_columns = self._parse_store_columns(sheet)

        # Extract date from filename pattern "Continuity Supplier Size Report DD-MM-YYYY.xlsx" or "DD_MM_YYYY.xlsx"
        # Example: "28-09-2025.xlsx" or "27_04_2025.xlsx" -> datetime(2025, 9, 28)
        file_date = None
        date_match = re.search(r'(\d{2})[-_](\d{2})[-_](\d{4})\.xlsx$', file_path, re.IGNORECASE)
        if date_match:
            day, month, year = date_match.groups()
            try:
                file_date = datetime(int(year), int(month), int(day))
                print(f"[Liberty] Extracted date from filename: {file_date.date()}")
            except ValueError as e:
                print(f"[Liberty] Invalid date in filename ({day}/{month}/{year}): {e}")
                file_date = None
        else:
            print(f"[Liberty] Could not extract date from filename: {file_path}")

        # Read row 3 as headers
        headers = []
        for cell in sheet[3]:  # Row 3 (1-indexed)
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                headers.append("")

        print(f"[Liberty] Found {len(headers)} columns in row 3")
        print(f"[Liberty] Detected {len(store_columns)} stores with data")

        # Process data rows starting from row 4
        # Liberty uses 3-row pattern: description row + blank row + Liberty ID/data row
        rows = []
        all_rows = list(sheet.iter_rows(min_row=4, values_only=True))

        i = 0
        while i < len(all_rows):
            description_row = all_rows[i]

            # Skip empty rows
            if not any(description_row):
                i += 1
                continue

            # Check if this is a product description row (has text in column F)
            description = description_row[5] if len(description_row) > 5 else None  # Column F (0-indexed: 5)

            # Check if row i+2 exists and has Liberty identifier
            if i + 2 < len(all_rows):
                data_row = all_rows[i + 2]
                liberty_name = data_row[5] if len(data_row) > 5 else None  # Column F (0-indexed: 5)

                # Liberty identifier format: "000834429 | 98-NO COLOUR Total"
                # Must have "|" and end with "Total"
                if liberty_name and isinstance(liberty_name, str) and '|' in liberty_name and liberty_name.endswith('Total'):
                    # This is a valid 3-row product pattern

                    # Create MULTIPLE records - one per store with data
                    for store_id, col_info in store_columns.items():
                        qty_col = col_info.get('qty_col')
                        sales_col = col_info.get('sales_col')

                        if qty_col is None or sales_col is None:
                            print(f"[Liberty] WARNING: Store '{store_id}' missing columns (qty_col: {qty_col}, sales_col: {sales_col}) - skipping")
                            continue

                        # Check if this store has data (non-zero quantity or sales)
                        qty_value = data_row[qty_col] if qty_col < len(data_row) else None
                        sales_value = data_row[sales_col] if sales_col < len(data_row) else None

                        # Skip if no data for this store
                        if not qty_value and not sales_value:
                            continue

                        # Create a record for this store with Liberty identifier
                        store_row = {
                            'liberty_name': liberty_name,  # Liberty identifier for product lookup
                            'Item': description,  # Keep description for reference
                            'Sales Qty Un': qty_value,
                            'Sales Inc VAT £ ': sales_value,
                            'store_identifier': store_id,  # Add store identifier
                            '_file_date': file_date  # Add extracted date (None if not found)
                        }

                        rows.append(store_row)

                    # Skip all 3 rows (description + blank + data)
                    i += 3
                else:
                    # Not a valid 3-row pattern, skip this row
                    i += 1
            else:
                # Not enough rows left for 3-row pattern, skip
                i += 1

        workbook.close()
        print(f"[Liberty] Extracted {len(rows)} sales records across {len(store_columns)} stores")
        return rows

    def transform_row(
        self,
        raw_row: Dict[str, Any],
        batch_id: str
    ) -> Optional[Dict[str, Any]]:
        """
        Transform Liberty row to sales_unified schema

        Handles:
        1. Liberty identifier lookup: "000834429 | 98-NO COLOUR Total" → products table
        2. Returns in parentheses: (123) → -123
        3. GBP → EUR conversion
        4. Store identification from column position
        5. Monthly "Actual" sales data extraction (not YTD)

        Liberty data structure:
        - "liberty_name" contains Liberty identifier (e.g., "000834429 | 98-NO COLOUR Total")
        - Lookup EAN and functional_name from products table
        - "Actual" columns have monthly sales data
        - Multiple store columns (Flagship, Internet, etc.)
        """
        # Start with base row
        transformed = self._create_base_row(batch_id)

        # Extract Liberty identifier from raw_row
        # Format: "000834429 | 98-NO COLOUR Total"
        liberty_name = raw_row.get("liberty_name")

        if not liberty_name:
            # Skip row - no Liberty identifier found
            return None

        # Lookup product in pre-loaded products table
        product = self.liberty_products.get(liberty_name)

        if product:
            # Match found - use EAN and functional_name from products table
            transformed["product_ean"] = product["ean"]
            transformed["functional_name"] = product["functional_name"]
        else:
            # No match found - still insert record with Liberty name as temporary identifier
            # Track as unmatched for reporting but don't discard the sales data
            self.unmatched_liberty_names.append(liberty_name)
            # Use Liberty name as temporary product identifier (database allows this per schema comment)
            transformed["product_ean"] = liberty_name
            transformed["functional_name"] = liberty_name

        # Store Liberty identifier for reference
        transformed["product_name_raw"] = liberty_name

        # Extract quantity - Liberty has no single "quantity" column
        # Instead, quantity is in store-specific columns like "Sales Qty Un"
        # For now, we'll look for YTD quantity (columns after "Actual")

        # Try to find YTD Sales Qty Un column
        qty_value = None
        sales_value = None

        # Liberty file structure from row 3:
        # ...YTD columns are: "Sales Qty Un", "Sales Inc VAT £ " (with trailing space)

        # Check for various quantity column names
        for col_name in ["Sales Qty Un", "Sold", "Quantity"]:
            if col_name in raw_row and raw_row[col_name] is not None:
                qty_value = raw_row[col_name]
                break

        if qty_value is None or qty_value == '':
            # No quantity data - skip this row (might be a header or summary row)
            return None

        try:
            quantity = self._to_int(qty_value, "Sales Qty Un")
            if quantity == 0:
                # Zero quantity - skip
                return None

            transformed["quantity"] = quantity

            # If quantity is negative (returns), mark as return
            if quantity < 0:
                transformed["is_return"] = True
            else:
                transformed["is_return"] = False

        except ValueError as e:
            raise ValueError(f"Invalid quantity: {e}")

        # Extract sales amount in GBP
        # Try different column name variations
        for col_name in ["Sales Inc VAT £ ", "Sales Inc VAT £", "Value", "Sales"]:
            if col_name in raw_row and raw_row[col_name] is not None and raw_row[col_name] != '':
                sales_value = raw_row[col_name]
                break

        if sales_value is None:
            raise ValueError("Missing sales amount")

        try:
            sales_gbp = self._to_float(sales_value, "Sales Inc VAT")

            # Store local currency amount
            transformed["sales_local_currency"] = sales_gbp

            # Convert to EUR
            transformed["sales_eur"] = self._convert_currency(sales_gbp, "GBP")

        except ValueError as e:
            raise ValueError(f"Invalid sales amount: {e}")

        # Extract date from filename (parsed in extract_rows and stored in raw_row)
        # Fallback to current date if not available
        file_date = raw_row.get("_file_date")
        if file_date:
            sale_date = file_date
        else:
            sale_date = datetime.utcnow()
            print(f"[Liberty] WARNING: No file date found, using current date")

        transformed["sale_date"] = sale_date.date().isoformat()
        transformed["year"] = sale_date.year
        transformed["month"] = sale_date.month
        transformed["quarter"] = self._calculate_quarter(sale_date.month)

        # Store identification: Use the store_identifier from raw_row
        # This was set during extract_rows based on which store column had the data
        store_identifier = raw_row.get("store_identifier", "flagship")
        transformed["store_identifier"] = store_identifier

        # Customer ID - Liberty is B2B reseller data, no customer information
        transformed["customer_id"] = None

        # Geography - Liberty is always UK
        transformed["country"] = "UK"

        # City and sales_channel based on store type
        # For Liberty, sales_channel represents distribution channel (not business model)
        if store_identifier.lower() in ["online", "internet"]:
            # Online/e-commerce sales
            transformed["city"] = "online"
            transformed["sales_channel"] = "online"
        else:
            # Physical store sales (flagship, etc.)
            transformed["city"] = "London"
            transformed["sales_channel"] = "retail"

        # Set upload_id for FK to uploads table
        # BIBBI schema uses upload_id (FK to uploads.id)
        # The batch_id parameter passed to transform_row is the upload UUID
        transformed["upload_id"] = batch_id

        return transformed

    def process_file(self, file_path: str) -> 'ProcessingResult':
        """Override base process_file to add unmatched Liberty names reporting"""
        # Call parent process_file
        result = super().process_file(file_path)

        # Print unmatched Liberty names report
        if self.unmatched_liberty_names:
            print(f"\n[Liberty] ========== UNMATCHED LIBERTY NAMES ==========")
            print(f"[Liberty] Total unmatched: {len(self.unmatched_liberty_names)}")
            print(f"[Liberty] Unique unmatched: {len(set(self.unmatched_liberty_names))}")
            print(f"[Liberty] ")
            print(f"[Liberty] Liberty names not found in products table:")
            for liberty_name in sorted(set(self.unmatched_liberty_names)):
                print(f"[Liberty]   - {liberty_name}")
            print(f"[Liberty] ==============================================\n")

        return result


def get_liberty_processor(reseller_id: str, supabase_client: 'Client') -> LibertyProcessor:
    """
    Factory function to create Liberty processor

    Args:
        reseller_id: UUID of Liberty reseller from resellers table
        supabase_client: Supabase client for product lookups

    Returns:
        LibertyProcessor instance
    """
    return LibertyProcessor(reseller_id, supabase_client)

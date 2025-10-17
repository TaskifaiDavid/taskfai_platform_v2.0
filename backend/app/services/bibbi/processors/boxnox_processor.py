"""
Boxnox Processor (BIBBI Version)

Handles Boxnox reseller data:
1. POS column = store identifier
2. Living document: Monthly additions to same file
3. Currency: EUR (no conversion needed)
4. Sheet name: "Sell Out by EAN"

Based on: backend/BIBBI/Resellers/resellers_info.md
"""

from typing import List, Dict, Any, Optional
from datetime import datetime
import openpyxl

from .base import BibbiBseProcessor


class BoxnoxProcessor(BibbiBseProcessor):
    """Process Boxnox Excel files"""

    VENDOR_NAME = "boxnox"
    CURRENCY = "EUR"
    TARGET_SHEET = "Sell Out by EAN"

    COLUMN_MAPPING = {
        "Product EAN": "product_ean",
        "Functional Name": "functional_name",
        "Sold Qty": "quantity",
        "Sales Amount (EUR)": "sales_eur",
        "Sales Amount": "sales_eur",
        "POS": "pos",
        "Month": "month",
        "Year": "year"
    }

    def get_vendor_name(self) -> str:
        return self.VENDOR_NAME

    def get_currency(self) -> str:
        return self.CURRENCY

    def extract_stores(self, file_path: str) -> List[Dict[str, Any]]:
        """Extract stores from POS column"""
        stores = []
        seen_stores = set()

        try:
            workbook = self._load_workbook(file_path)

            # Find correct sheet
            sheet = None
            if self.TARGET_SHEET in workbook.sheetnames:
                sheet = workbook[self.TARGET_SHEET]
            else:
                sheet = workbook[workbook.sheetnames[0]]

            headers = self._get_sheet_headers(sheet)

            # Find POS column
            pos_col_idx = None
            for idx, header in enumerate(headers):
                if "POS" in header.upper():
                    pos_col_idx = idx
                    break

            if pos_col_idx is not None:
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue

                    pos_value = row[pos_col_idx] if pos_col_idx < len(row) else None
                    if pos_value:
                        pos_str = str(pos_value).strip()
                        if pos_str and pos_str not in seen_stores:
                            seen_stores.add(pos_str)

                            # Determine if online
                            is_online = any(kw in pos_str.lower() for kw in ["online", "web", "e-shop"])

                            stores.append({
                                "store_identifier": pos_str.lower().replace(' ', '_'),
                                "store_name": f"Boxnox {pos_str}",
                                "store_type": "online" if is_online else "physical",
                                "reseller_id": self.reseller_id
                            })

            workbook.close()

            if not stores:
                # Fallback
                stores = [{
                    "store_identifier": "boxnox_main",
                    "store_name": "Boxnox Main",
                    "store_type": "physical",
                    "reseller_id": self.reseller_id
                }]

        except Exception as e:
            print(f"[Boxnox] Error extracting stores: {e}")
            stores = [{
                "store_identifier": "boxnox_main",
                "store_name": "Boxnox Main",
                "store_type": "physical",
                "reseller_id": self.reseller_id
            }]

        return stores

    def extract_rows(self, file_path: str) -> List[Dict[str, Any]]:
        """Extract rows from Boxnox file"""
        workbook = self._load_workbook(file_path)

        # Find correct sheet
        if self.TARGET_SHEET in workbook.sheetnames:
            sheet = workbook[self.TARGET_SHEET]
        else:
            sheet = workbook[workbook.sheetnames[0]]

        headers = self._get_sheet_headers(sheet)

        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            row_dict = {}
            for idx, header in enumerate(headers):
                if idx < len(row):
                    row_dict[header] = row[idx]

            rows.append(row_dict)

        workbook.close()
        return rows

    def transform_row(self, raw_row: Dict[str, Any], batch_id: str) -> Optional[Dict[str, Any]]:
        """Transform Boxnox row to sales_unified schema"""
        transformed = self._create_base_row(batch_id)

        # EAN
        ean_value = raw_row.get("Product EAN")
        if not ean_value:
            raise ValueError("Missing Product EAN")

        transformed["product_id"] = self._validate_ean(ean_value)

        # Quantity
        qty_value = raw_row.get("Sold Qty")
        if qty_value is None:
            raise ValueError("Missing Sold Qty")

        quantity = self._to_int(qty_value, "Sold Qty")
        if quantity <= 0:
            raise ValueError(f"Invalid quantity: {quantity}")

        transformed["quantity"] = quantity
        transformed["is_return"] = False

        # Sales amount (already in EUR)
        sales_value = raw_row.get("Sales Amount (EUR)") or raw_row.get("Sales Amount")
        if sales_value is None:
            raise ValueError("Missing Sales Amount")

        sales_eur = self._to_float(sales_value, "Sales Amount")
        transformed["sales_local_currency"] = sales_eur
        transformed["sales_eur"] = sales_eur

        # Date
        month_value = raw_row.get("Month")
        year_value = raw_row.get("Year")

        if month_value and year_value:
            month = self._to_int(month_value, "Month")
            year = self._to_int(year_value, "Year")

            if month < 1 or month > 12:
                raise ValueError(f"Invalid month: {month}")

            transformed["sale_date"] = datetime(year, month, 1).date().isoformat()
            transformed["year"] = year
            transformed["month"] = month
            transformed["quarter"] = self._calculate_quarter(month)
        else:
            now = datetime.utcnow()
            transformed["sale_date"] = now.date().isoformat()
            transformed["year"] = now.year
            transformed["month"] = now.month
            transformed["quarter"] = self._calculate_quarter(now.month)

        # Store (POS)
        pos_value = raw_row.get("POS")
        if pos_value:
            transformed["store_identifier"] = str(pos_value).strip().lower().replace(' ', '_')
        else:
            transformed["store_identifier"] = "boxnox_main"

        return transformed


def get_boxnox_processor(reseller_id: str) -> BoxnoxProcessor:
    """Factory function for Boxnox processor"""
    return BoxnoxProcessor(reseller_id)

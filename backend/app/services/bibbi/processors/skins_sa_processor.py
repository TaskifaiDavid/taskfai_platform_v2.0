"""
Skins SA Processor (BIBBI Version)

Handles Skins South Africa reseller data with special cases:
1. Column A store codes: "ON" = online, others = physical store names
2. Living document: Same file updated monthly (duplicate detection needed)
3. Dual sales columns: netsales and exvatnetsales (use exvatnetsales)
4. EAN column: Called "Stockcode" in file
5. Currency: ZAR → EUR conversion

Based on: backend/BIBBI/Resellers/resellers_info.md
"""

from typing import List, Dict, Any, Optional
from datetime import datetime
import openpyxl

from .base import BibbiBseProcessor


class SkinsSAProcessor(BibbiBseProcessor):
    """Process Skins South Africa Excel files with ZAR to EUR conversion"""

    VENDOR_NAME = "skins_sa"
    CURRENCY = "ZAR"
    ZAR_TO_EUR_RATE = 0.05

    # Column mapping from Skins SA format to internal names
    COLUMN_MAPPING = {
        "OrderDate": "order_date",
        "Stockcode": "product_ean",
        "StockCode": "product_ean",  # Case variations
        "EAN": "product_ean",
        "Qty": "quantity",
        "Quantity": "quantity",
        "Amount": "sales_zar",
        "netsales": "netsales",
        "exvatnetsales": "exvatnetsales",  # Prefer this
        "Month": "month",
        "Year": "year"
    }

    def get_vendor_name(self) -> str:
        return self.VENDOR_NAME

    def get_currency(self) -> str:
        return self.CURRENCY

    def extract_stores(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Extract Skins SA store information

        CRITICAL: Column A contains store codes
        - "ON" = online store
        - All other values = physical store names

        Returns:
            List of store dictionaries
        """
        stores = []
        seen_stores = set()

        try:
            workbook = self._load_workbook(file_path)
            sheet = workbook[workbook.sheetnames[0]]

            # Column A (index 0) contains store codes
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                # Get Column A value (store code)
                store_code = row[0] if len(row) > 0 else None
                if not store_code:
                    continue

                store_str = str(store_code).strip()
                if not store_str or store_str in seen_stores:
                    continue

                seen_stores.add(store_str)

                # Determine store type
                store_identifier = store_str.lower().replace(' ', '_')
                is_online = store_str.upper() == "ON"

                stores.append({
                    "store_identifier": store_identifier,
                    "store_name": f"Skins {store_str}" if not is_online else "Skins Online",
                    "store_type": "online" if is_online else "physical",
                    "reseller_id": self.reseller_id,
                    "country": "South Africa"
                })

            workbook.close()

            if not stores:
                # Fallback: Create default online store
                stores = [{
                    "store_identifier": "on",
                    "store_name": "Skins Online",
                    "store_type": "online",
                    "reseller_id": self.reseller_id,
                    "country": "South Africa"
                }]

        except Exception as e:
            print(f"[SkinsSA] Error extracting stores: {e}")
            # Fallback
            stores = [{
                "store_identifier": "on",
                "store_name": "Skins Online",
                "store_type": "online",
                "reseller_id": self.reseller_id,
                "country": "South Africa"
            }]

        return stores

    def extract_rows(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Extract rows from Skins SA Excel file

        CRITICAL: Must include Column A (store code) in extraction
        """
        workbook = self._load_workbook(file_path)
        sheet = workbook[workbook.sheetnames[0]]
        headers = self._get_sheet_headers(sheet)

        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            row_dict = {}

            # Always capture Column A as store_code
            row_dict["_store_code_column_a"] = row[0] if len(row) > 0 else None

            # Map other columns by header
            for idx, header in enumerate(headers):
                if idx < len(row):
                    row_dict[header] = row[idx]

            rows.append(row_dict)

        workbook.close()
        return rows

    def transform_row(
        self,
        raw_row: Dict[str, Any],
        batch_id: str
    ) -> Optional[Dict[str, Any]]:
        """
        Transform Skins SA row to sales_unified schema

        Handles:
        1. Column A store code parsing
        2. Stockcode → EAN
        3. exvatnetsales preferred over netsales
        4. ZAR → EUR conversion
        5. Date parsing (multiple formats)
        """
        # Start with base row
        transformed = self._create_base_row(batch_id)

        # Extract EAN from Stockcode column
        ean_value = raw_row.get("Stockcode") or raw_row.get("StockCode") or raw_row.get("EAN")
        if not ean_value:
            raise ValueError("Missing Stockcode/EAN")

        try:
            transformed["product_id"] = self._validate_ean(ean_value)
        except ValueError as e:
            raise ValueError(f"Invalid EAN: {e}")

        # Extract quantity
        qty_value = raw_row.get("Qty") or raw_row.get("Quantity")
        if qty_value is None:
            raise ValueError("Missing Qty")

        try:
            quantity = self._to_int(qty_value, "Qty")
            if quantity <= 0:
                raise ValueError(f"Invalid quantity: {quantity}")

            transformed["quantity"] = quantity
            transformed["is_return"] = False

        except ValueError as e:
            raise ValueError(f"Invalid quantity: {e}")

        # Extract sales amount (prefer exvatnetsales over netsales)
        sales_value = raw_row.get("exvatnetsales") or raw_row.get("Amount") or raw_row.get("netsales")
        if sales_value is None:
            raise ValueError("Missing sales amount")

        try:
            sales_zar = self._to_float(sales_value, "Amount")

            # Store local currency amount
            transformed["sales_local_currency"] = sales_zar

            # Convert to EUR
            transformed["sales_eur"] = self._convert_currency(sales_zar, "ZAR")

        except ValueError as e:
            raise ValueError(f"Invalid sales amount: {e}")

        # Extract date information
        # Try OrderDate column first
        date_value = raw_row.get("OrderDate")
        if date_value:
            try:
                sale_date = self._validate_date(date_value)
                transformed["sale_date"] = sale_date.date().isoformat()
                transformed["year"] = sale_date.year
                transformed["month"] = sale_date.month
                transformed["quarter"] = self._calculate_quarter(sale_date.month)

            except ValueError as e:
                # Fall back to Month/Year if OrderDate parsing fails
                month_value = raw_row.get("Month")
                year_value = raw_row.get("Year")

                if month_value and year_value:
                    month = self._to_int(month_value, "Month")
                    year = self._to_int(year_value, "Year")

                    if month < 1 or month > 12:
                        raise ValueError(f"Invalid month: {month}")

                    if year < 2000 or year > 2100:
                        raise ValueError(f"Invalid year: {year}")

                    transformed["sale_date"] = datetime(year, month, 1).date().isoformat()
                    transformed["year"] = year
                    transformed["month"] = month
                    transformed["quarter"] = self._calculate_quarter(month)
                else:
                    raise ValueError(f"Invalid date: {e}")
        else:
            # Use Month/Year columns
            month_value = raw_row.get("Month")
            year_value = raw_row.get("Year")

            if month_value and year_value:
                try:
                    month = self._to_int(month_value, "Month")
                    year = self._to_int(year_value, "Year")

                    if month < 1 or month > 12:
                        raise ValueError(f"Invalid month: {month}")

                    if year < 2000 or year > 2100:
                        raise ValueError(f"Invalid year: {year}")

                    transformed["sale_date"] = datetime(year, month, 1).date().isoformat()
                    transformed["year"] = year
                    transformed["month"] = month
                    transformed["quarter"] = self._calculate_quarter(month)

                except ValueError as e:
                    raise ValueError(f"Invalid date: {e}")
            else:
                # Use current date if not provided
                now = datetime.utcnow()
                transformed["sale_date"] = now.date().isoformat()
                transformed["year"] = now.year
                transformed["month"] = now.month
                transformed["quarter"] = self._calculate_quarter(now.month)

        # Extract store from Column A
        store_code = raw_row.get("_store_code_column_a")
        if store_code:
            store_str = str(store_code).strip().lower().replace(' ', '_')
            transformed["store_identifier"] = store_str
        else:
            # Default to online if not specified
            transformed["store_identifier"] = "on"

        return transformed


def get_skins_sa_processor(reseller_id: str) -> SkinsSAProcessor:
    """
    Factory function to create Skins SA processor

    Args:
        reseller_id: UUID of Skins SA reseller from resellers table

    Returns:
        SkinsSAProcessor instance
    """
    return SkinsSAProcessor(reseller_id)

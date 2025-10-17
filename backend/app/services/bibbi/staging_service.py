"""
BIBBI Staging Service

Manages the staging layer for uploaded reseller files.
Raw file data is stored in sales_staging table before processing.

Pipeline Stage: Upload → **STAGING** → Validation → Unified
"""

import json
import uuid
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, Optional
import openpyxl

from app.core.bibbi import BibbιDB, BIBBI_TENANT_ID


class BibbιStagingService:
    """
    Service for staging uploaded BIBBI reseller files

    Responsibilities:
    - Insert raw file metadata into sales_staging table
    - Store file information (filename, size, hash) as JSONB
    - Extract basic file structure (sheet names, row counts)
    - Link staging record to upload record
    """

    def __init__(self, bibbi_db: BibbιDB):
        """
        Initialize staging service

        Args:
            bibbi_db: BIBBI-specific Supabase client
        """
        self.db = bibbi_db

    def stage_upload(
        self,
        upload_id: str,
        file_path: str,
        filename: str,
        file_size: int,
        file_hash: str
    ) -> str:
        """
        Stage an uploaded file for processing

        Creates a staging record with raw file metadata before processing begins.

        Args:
            upload_id: Unique upload identifier
            file_path: Path to uploaded file on disk
            filename: Original filename
            file_size: File size in bytes
            file_hash: SHA256 hash of file content

        Returns:
            staging_id: Unique staging record identifier

        Raises:
            Exception: If staging fails
        """
        # Generate unique staging ID
        staging_id = str(uuid.uuid4())

        # Extract file metadata
        file_metadata = self._extract_file_metadata(file_path, filename)

        # Create staging record
        staging_data = {
            "staging_id": staging_id,
            "upload_id": upload_id,
            "file_path": file_path,
            "filename": filename,
            "file_size": file_size,
            "file_hash": file_hash,
            "raw_data": json.dumps(file_metadata),  # Store as JSONB
            "staging_status": "pending",
            "created_at": datetime.utcnow().isoformat(),
            # tenant_id automatically added by BibbιSupabaseClient
        }

        try:
            result = self.db.table("sales_staging").insert(staging_data).execute()

            if not result.data:
                raise Exception("Failed to create staging record")

            print(f"[BibbιStaging] Created staging record: {staging_id}")
            return staging_id

        except Exception as e:
            print(f"[BibbιStaging] Error creating staging record: {e}")
            raise Exception(f"Failed to stage upload: {str(e)}")

    def update_staging_status(
        self,
        staging_id: str,
        status: str,
        validation_errors: Optional[Dict] = None
    ) -> None:
        """
        Update staging record status

        Args:
            staging_id: Unique staging identifier
            status: New status (pending, processing, validated, failed)
            validation_errors: Optional validation errors (dict)

        Raises:
            Exception: If update fails
        """
        update_data = {
            "staging_status": status,
            "updated_at": datetime.utcnow().isoformat()
        }

        if validation_errors:
            update_data["validation_errors"] = json.dumps(validation_errors)

        try:
            self.db.table("sales_staging")\
                .update(update_data)\
                .eq("staging_id", staging_id)\
                .execute()

            print(f"[BibbιStaging] Updated staging status: {staging_id} → {status}")

        except Exception as e:
            print(f"[BibbιStaging] Error updating staging status: {e}")
            raise Exception(f"Failed to update staging status: {str(e)}")

    def update_validation_results(
        self,
        staging_id: str,
        validation_result: Any  # ValidationResult from validation_service
    ) -> None:
        """
        Update staging record with validation results

        Stores comprehensive validation outcome including:
        - Validation status (validated or validation_failed)
        - Row-level error details
        - Validation timestamp
        - Success/failure statistics

        Args:
            staging_id: Unique staging identifier
            validation_result: ValidationResult object with validation outcome

        Raises:
            Exception: If update fails
        """
        # Determine status based on validation outcome
        if validation_result.invalid_rows == 0:
            status = "validated"
        else:
            status = "validation_failed"

        # Format validation errors for JSONB storage
        validation_errors = {
            "total_rows": validation_result.total_rows,
            "valid_rows": validation_result.valid_rows,
            "invalid_rows": validation_result.invalid_rows,
            "validation_success_rate": round(
                validation_result.valid_rows / validation_result.total_rows * 100, 2
            ) if validation_result.total_rows > 0 else 0,
            "errors": validation_result.errors  # List of error dicts with row_number, error_type, error_message
        }

        update_data = {
            "staging_status": status,
            "validation_status": status,  # Separate field for validation status
            "validation_errors": json.dumps(validation_errors),
            "validated_at": datetime.utcnow().isoformat(),
            "updated_at": datetime.utcnow().isoformat()
        }

        try:
            self.db.table("sales_staging")\
                .update(update_data)\
                .eq("staging_id", staging_id)\
                .execute()

            print(f"[BibbιStaging] Updated validation results: {staging_id} → {status}")
            print(f"[BibbιStaging] Valid: {validation_result.valid_rows}/{validation_result.total_rows} "
                  f"({validation_errors['validation_success_rate']}%)")

        except Exception as e:
            print(f"[BibbιStaging] Error updating validation results: {e}")
            raise Exception(f"Failed to update validation results: {str(e)}")

    def get_staging_record(self, staging_id: str) -> Optional[Dict]:
        """
        Get staging record by ID

        Args:
            staging_id: Unique staging identifier

        Returns:
            Staging record dict or None if not found
        """
        try:
            result = self.db.table("sales_staging")\
                .select("*")\
                .eq("staging_id", staging_id)\
                .execute()

            if result.data and len(result.data) > 0:
                return result.data[0]

            return None

        except Exception as e:
            print(f"[BibbιStaging] Error getting staging record: {e}")
            return None

    def get_validation_errors(self, staging_id: str) -> Optional[Dict]:
        """
        Get validation errors for a staging record

        Retrieves and parses validation errors from JSONB storage.

        Args:
            staging_id: Unique staging identifier

        Returns:
            Dictionary with validation errors or None if not found
        """
        try:
            result = self.db.table("sales_staging")\
                .select("validation_errors, validation_status, validated_at")\
                .eq("staging_id", staging_id)\
                .execute()

            if result.data and len(result.data) > 0:
                record = result.data[0]

                if record.get("validation_errors"):
                    # Parse JSONB back to dict
                    validation_errors = json.loads(record["validation_errors"]) \
                        if isinstance(record["validation_errors"], str) \
                        else record["validation_errors"]

                    return {
                        "validation_status": record.get("validation_status"),
                        "validated_at": record.get("validated_at"),
                        **validation_errors
                    }

            return None

        except Exception as e:
            print(f"[BibbιStaging] Error getting validation errors: {e}")
            return None

    def _extract_file_metadata(self, file_path: str, filename: str) -> Dict[str, Any]:
        """
        Extract metadata from Excel file

        Reads file structure without processing full data:
        - Sheet names
        - Row counts per sheet
        - Column headers per sheet
        - First row sample (for debugging)

        Args:
            file_path: Path to Excel file
            filename: Original filename

        Returns:
            Dictionary with file metadata
        """
        metadata = {
            "filename": filename,
            "file_extension": Path(filename).suffix,
            "upload_timestamp": datetime.utcnow().isoformat(),
            "sheets": []
        }

        try:
            # Load workbook in read-only mode for performance
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Get column headers (first row)
                headers = []
                first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if first_row:
                    headers = [str(cell) if cell is not None else "" for cell in first_row]

                # Count rows (approximate - read-only mode doesn't provide accurate count)
                row_count = sum(1 for _ in sheet.iter_rows())

                # Get sample first data row (skip header)
                sample_row = []
                data_rows = sheet.iter_rows(min_row=2, max_row=2, values_only=True)
                first_data_row = next(data_rows, None)
                if first_data_row:
                    sample_row = [str(cell)[:50] if cell is not None else "" for cell in first_data_row]

                sheet_metadata = {
                    "sheet_name": sheet_name,
                    "row_count": row_count,
                    "column_count": len(headers),
                    "headers": headers,
                    "sample_row": sample_row
                }

                metadata["sheets"].append(sheet_metadata)

            workbook.close()

            metadata["total_sheets"] = len(metadata["sheets"])
            metadata["extraction_success"] = True

        except Exception as e:
            metadata["extraction_success"] = False
            metadata["extraction_error"] = str(e)
            print(f"[BibbιStaging] Error extracting metadata: {e}")

        return metadata

    def link_staging_to_upload(self, upload_id: str, staging_id: str) -> None:
        """
        Link staging record to upload record

        Updates upload record with staging_id for tracking.

        Args:
            upload_id: Unique upload identifier
            staging_id: Unique staging identifier

        Raises:
            Exception: If linking fails
        """
        try:
            self.db.table("uploads")\
                .update({
                    "staging_id": staging_id,
                    "upload_status": "staged",
                    "updated_at": datetime.utcnow().isoformat()
                })\
                .eq("upload_id", upload_id)\
                .execute()

            print(f"[BibbιStaging] Linked staging to upload: {upload_id} → {staging_id}")

        except Exception as e:
            print(f"[BibbιStaging] Error linking staging to upload: {e}")
            raise Exception(f"Failed to link staging to upload: {str(e)}")


def get_staging_service(bibbi_db: BibbιDB) -> BibbιStagingService:
    """
    Factory function to create staging service

    Args:
        bibbi_db: BIBBI-specific Supabase client

    Returns:
        BibbιStagingService instance
    """
    return BibbιStagingService(bibbi_db)

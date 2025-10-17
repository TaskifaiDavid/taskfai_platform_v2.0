"""
BIBBI Error Report Service

Generates Excel validation error reports for failed uploads.

Pipeline Stage: Validation → **ERROR REPORTING** → User Download

Features:
- Summary sheet with validation statistics
- Error details sheet with row-level errors
- User-friendly formatting and highlighting
"""

import uuid
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, Optional, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.core.bibbi import BibbιDB, BIBBI_TENANT_ID


class BibbιErrorReportService:
    """
    Service for generating Excel validation error reports

    Responsibilities:
    - Generate Excel files with validation error details
    - Format reports with summary and error sheets
    - Apply styling for readability
    - Save reports to disk for download
    """

    # Color scheme for Excel styling
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ERROR_FILL = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    SUCCESS_FILL = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BOLD_FONT = Font(bold=True, size=11)

    def __init__(self, bibbi_db: BibbιDB, report_dir: str = "/tmp/bibbi_reports"):
        """
        Initialize error report service

        Args:
            bibbi_db: BIBBI-specific Supabase client
            report_dir: Directory to save generated reports
        """
        self.db = bibbi_db
        self.report_dir = Path(report_dir)
        self.report_dir.mkdir(parents=True, exist_ok=True)

    def generate_error_report(
        self,
        staging_id: str,
        validation_errors: Dict[str, Any],
        filename: str = None
    ) -> str:
        """
        Generate Excel error report from validation errors

        Creates a multi-sheet Excel workbook with:
        - Summary sheet: Validation statistics and overview
        - Error Details sheet: Row-level error information

        Args:
            staging_id: Unique staging identifier
            validation_errors: Validation error data from staging record
            filename: Optional filename for report (default: auto-generated)

        Returns:
            report_path: Path to generated Excel report

        Example validation_errors format:
        {
            "total_rows": 100,
            "valid_rows": 85,
            "invalid_rows": 15,
            "validation_success_rate": 85.0,
            "errors": [
                {
                    "row_number": 12,
                    "error_type": "validation_error",
                    "error_message": "Missing required field: product_id",
                    "row_data": {...}
                }
            ]
        }
        """
        # Generate filename if not provided
        if filename is None:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"validation_errors_{staging_id[:8]}_{timestamp}.xlsx"

        report_path = self.report_dir / filename

        # Create workbook
        workbook = openpyxl.Workbook()

        # Remove default sheet
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

        # Create sheets
        self._create_summary_sheet(workbook, staging_id, validation_errors)
        self._create_error_details_sheet(workbook, validation_errors.get("errors", []))

        # Save workbook
        workbook.save(report_path)
        print(f"[BibbιErrorReport] Generated report: {report_path}")

        return str(report_path)

    def _create_summary_sheet(
        self,
        workbook: openpyxl.Workbook,
        staging_id: str,
        validation_errors: Dict[str, Any]
    ) -> None:
        """
        Create summary sheet with validation statistics

        Args:
            workbook: Excel workbook
            staging_id: Staging record identifier
            validation_errors: Validation error data
        """
        sheet = workbook.create_sheet("Summary", 0)

        # Title
        sheet["A1"] = "BIBBI Validation Error Report"
        sheet["A1"].font = Font(bold=True, size=14)
        sheet.merge_cells("A1:B1")

        # Report metadata
        sheet["A3"] = "Report Details"
        sheet["A3"].font = self.BOLD_FONT

        metadata = [
            ("Staging ID:", staging_id),
            ("Generated At:", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")),
            ("Tenant:", BIBBI_TENANT_ID)
        ]

        row = 4
        for label, value in metadata:
            sheet[f"A{row}"] = label
            sheet[f"A{row}"].font = self.BOLD_FONT
            sheet[f"B{row}"] = value
            row += 1

        # Validation statistics
        sheet[f"A{row + 1}"] = "Validation Statistics"
        sheet[f"A{row + 1}"].font = self.BOLD_FONT

        total_rows = validation_errors.get("total_rows", 0)
        valid_rows = validation_errors.get("valid_rows", 0)
        invalid_rows = validation_errors.get("invalid_rows", 0)
        success_rate = validation_errors.get("validation_success_rate", 0)

        stats = [
            ("Total Rows:", total_rows),
            ("Valid Rows:", valid_rows),
            ("Invalid Rows:", invalid_rows),
            ("Success Rate:", f"{success_rate}%")
        ]

        row += 2
        for label, value in stats:
            sheet[f"A{row}"] = label
            sheet[f"A{row}"].font = self.BOLD_FONT
            sheet[f"B{row}"] = value

            # Color code success rate
            if label == "Success Rate:":
                if success_rate >= 95:
                    sheet[f"B{row}"].fill = self.SUCCESS_FILL
                elif success_rate < 80:
                    sheet[f"B{row}"].fill = self.ERROR_FILL

            row += 1

        # Error summary by type
        errors = validation_errors.get("errors", [])
        if errors:
            sheet[f"A{row + 1}"] = "Error Summary by Type"
            sheet[f"A{row + 1}"].font = self.BOLD_FONT

            error_types = {}
            for error in errors:
                error_type = error.get("error_type", "unknown")
                error_types[error_type] = error_types.get(error_type, 0) + 1

            row += 2
            sheet[f"A{row}"] = "Error Type"
            sheet[f"B{row}"] = "Count"
            sheet[f"A{row}"].font = self.HEADER_FONT
            sheet[f"B{row}"].font = self.HEADER_FONT
            sheet[f"A{row}"].fill = self.HEADER_FILL
            sheet[f"B{row}"].fill = self.HEADER_FILL

            row += 1
            for error_type, count in sorted(error_types.items()):
                sheet[f"A{row}"] = error_type
                sheet[f"B{row}"] = count
                sheet[f"A{row}"].fill = self.ERROR_FILL
                sheet[f"B{row}"].fill = self.ERROR_FILL
                row += 1

        # Column widths
        sheet.column_dimensions["A"].width = 25
        sheet.column_dimensions["B"].width = 40

    def _create_error_details_sheet(
        self,
        workbook: openpyxl.Workbook,
        errors: List[Dict[str, Any]]
    ) -> None:
        """
        Create error details sheet with row-level errors

        Args:
            workbook: Excel workbook
            errors: List of error dictionaries
        """
        sheet = workbook.create_sheet("Error Details")

        # Headers
        headers = ["Row Number", "Error Type", "Error Message", "Row Data (Sample)"]
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Error rows
        for row_idx, error in enumerate(errors, 2):
            # Row number
            sheet.cell(row=row_idx, column=1).value = error.get("row_number", "N/A")

            # Error type
            sheet.cell(row=row_idx, column=2).value = error.get("error_type", "unknown")

            # Error message
            error_message = error.get("error_message", "No details")
            sheet.cell(row=row_idx, column=3).value = error_message
            sheet.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True)

            # Row data (formatted)
            row_data = error.get("row_data", {})
            if row_data:
                # Format as key: value pairs (truncated)
                data_str = " | ".join([
                    f"{k}: {str(v)[:30]}" for k, v in list(row_data.items())[:3]
                ])
                sheet.cell(row=row_idx, column=4).value = data_str
                sheet.cell(row=row_idx, column=4).alignment = Alignment(wrap_text=True)

            # Highlight error rows
            for col in range(1, 5):
                sheet.cell(row=row_idx, column=col).fill = self.ERROR_FILL

        # Column widths
        sheet.column_dimensions["A"].width = 12
        sheet.column_dimensions["B"].width = 20
        sheet.column_dimensions["C"].width = 50
        sheet.column_dimensions["D"].width = 60

        # Freeze header row
        sheet.freeze_panes = "A2"

    def generate_report_from_staging(self, staging_id: str) -> Optional[str]:
        """
        Generate error report from staging record

        Convenience method that fetches validation errors from staging record.

        Args:
            staging_id: Unique staging identifier

        Returns:
            report_path: Path to generated report or None if no errors
        """
        try:
            # Get validation errors from staging record
            result = self.db.table("sales_staging")\
                .select("validation_errors, filename")\
                .eq("staging_id", staging_id)\
                .execute()

            if not result.data or len(result.data) == 0:
                print(f"[BibbιErrorReport] Staging record not found: {staging_id}")
                return None

            record = result.data[0]
            validation_errors = record.get("validation_errors")

            if not validation_errors:
                print(f"[BibbιErrorReport] No validation errors for: {staging_id}")
                return None

            # Parse JSONB if needed
            if isinstance(validation_errors, str):
                import json
                validation_errors = json.loads(validation_errors)

            # Check if there are actual errors
            if validation_errors.get("invalid_rows", 0) == 0:
                print(f"[BibbιErrorReport] No errors to report for: {staging_id}")
                return None

            # Generate report
            original_filename = record.get("filename", "unknown.xlsx")
            report_filename = f"errors_{Path(original_filename).stem}.xlsx"

            report_path = self.generate_error_report(
                staging_id=staging_id,
                validation_errors=validation_errors,
                filename=report_filename
            )

            return report_path

        except Exception as e:
            print(f"[BibbιErrorReport] Error generating report from staging: {e}")
            return None

    def cleanup_old_reports(self, days_old: int = 7) -> int:
        """
        Clean up old error reports

        Removes reports older than specified days.

        Args:
            days_old: Delete reports older than this many days

        Returns:
            count: Number of reports deleted
        """
        count = 0
        cutoff_time = datetime.utcnow().timestamp() - (days_old * 24 * 60 * 60)

        try:
            for report_file in self.report_dir.glob("*.xlsx"):
                if report_file.stat().st_mtime < cutoff_time:
                    report_file.unlink()
                    count += 1

            print(f"[BibbιErrorReport] Cleaned up {count} old reports")

        except Exception as e:
            print(f"[BibbιErrorReport] Error cleaning up reports: {e}")

        return count


def get_error_report_service(bibbi_db: BibbιDB, report_dir: str = "/tmp/bibbi_reports") -> BibbιErrorReportService:
    """
    Factory function to create error report service

    Args:
        bibbi_db: BIBBI-specific Supabase client
        report_dir: Directory to save generated reports

    Returns:
        BibbιErrorReportService instance
    """
    return BibbιErrorReportService(bibbi_db, report_dir)

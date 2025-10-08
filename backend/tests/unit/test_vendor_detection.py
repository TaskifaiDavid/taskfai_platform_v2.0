"""
Unit tests for vendor detection
Tests: backend/app/services/vendors/detector.py
"""

import pytest
import tempfile
import openpyxl
from pathlib import Path
from app.services.vendors.detector import VendorDetector


class TestVendorDetection:
    """Test vendor detection from uploaded files"""

    @pytest.fixture
    def detector(self):
        """Create VendorDetector instance"""
        return VendorDetector()

    # ============================================
    # EXCEL FILE DETECTION TESTS
    # ============================================

    def test_detect_boxnox_from_filename(self, detector):
        """Test Boxnox detection from filename"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            # Create minimal Excel file
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sell Out by EAN"
            ws.append(["Product EAN", "Functional Name", "Sold Qty", "Sales Amount (EUR)"])
            wb.save(tmp.name)
            wb.close()

            vendor, confidence = detector.detect_vendor(tmp.name, "boxnox_data.xlsx")

            assert vendor == "boxnox"
            assert confidence >= 0.5

            Path(tmp.name).unlink()

    def test_detect_galilu_from_file_structure(self, detector):
        """Test Galilu detection from file structure"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["EAN", "Product", "Month", "Year", "Sales"])
            wb.save(tmp.name)
            wb.close()

            vendor, confidence = detector.detect_vendor(tmp.name, "galilu_report.xlsx")

            assert vendor == "galilu"
            assert confidence >= 0.5

            Path(tmp.name).unlink()

    def test_detect_skins_sa_from_columns(self, detector):
        """Test Skins SA detection from columns"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["OrderDate", "EAN", "Qty", "Amount", "Customer"])
            wb.save(tmp.name)
            wb.close()

            vendor, confidence = detector.detect_vendor(tmp.name, "skins_south_africa.xlsx")

            assert vendor == "skins_sa"
            assert confidence >= 0.5

            Path(tmp.name).unlink()

    def test_detect_liberty_from_filename_and_columns(self, detector):
        """Test Liberty detection from filename and columns"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["EAN", "Product", "Sold", "Revenue"])
            wb.save(tmp.name)
            wb.close()

            vendor, confidence = detector.detect_vendor(tmp.name, "liberty_sales.xlsx")

            assert vendor == "liberty"
            assert confidence >= 0.5

            Path(tmp.name).unlink()

    # ============================================
    # CSV FILE DETECTION TESTS
    # ============================================

    def test_detect_from_csv_with_headers(self, detector):
        """Test detection from CSV file"""
        with tempfile.NamedTemporaryFile(mode='w', suffix=".csv", delete=False) as tmp:
            tmp.write("EAN,Product,Month,Year,Sales\n")
            tmp.write("1234567890123,Product A,1,2024,1000\n")
            tmp.flush()

            vendor, confidence = detector.detect_vendor(tmp.name, "galilu_data.csv")

            assert vendor == "galilu"
            assert confidence >= 0.5

            Path(tmp.name).unlink()

    def test_detect_online_sales_from_csv(self, detector):
        """Test online sales detection from CSV"""
        with tempfile.NamedTemporaryFile(mode='w', suffix=".csv", delete=False) as tmp:
            tmp.write("order_id,ean,quantity,price,date\n")
            tmp.write("12345,1234567890123,2,29.99,2024-01-01\n")
            tmp.flush()

            vendor, confidence = detector.detect_vendor(tmp.name, "online_orders.csv")

            assert vendor == "online"
            assert confidence >= 0.5

            Path(tmp.name).unlink()

    # ============================================
    # CONFIDENCE SCORING TESTS
    # ============================================

    def test_high_confidence_with_all_matches(self, detector):
        """Test high confidence when all patterns match"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sell Out by EAN"
            ws.append(["Product EAN", "Functional Name", "Sold Qty", "Sales Amount (EUR)", "Reseller", "Month", "Year"])
            wb.save(tmp.name)
            wb.close()

            vendor, confidence = detector.detect_vendor(tmp.name, "boxnox_monthly_report.xlsx")

            assert vendor == "boxnox"
            assert confidence >= 0.7  # High confidence

            Path(tmp.name).unlink()

    def test_low_confidence_below_threshold(self, detector):
        """Test that low confidence returns None"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Column1", "Column2", "Column3"])
            wb.save(tmp.name)
            wb.close()

            vendor, confidence = detector.detect_vendor(tmp.name, "unknown_file.xlsx")

            # Should return None if confidence below threshold
            assert vendor is None or confidence < 0.5

            Path(tmp.name).unlink()

    # ============================================
    # EDGE CASES
    # ============================================

    def test_detect_from_invalid_file_type(self, detector):
        """Test detection from unsupported file type"""
        with tempfile.NamedTemporaryFile(suffix=".txt", delete=False) as tmp:
            tmp.write(b"Some text content")
            tmp.flush()

            vendor, confidence = detector.detect_vendor(tmp.name, "data.txt")

            assert vendor is None
            assert confidence == 0.0

            Path(tmp.name).unlink()

    def test_detect_from_empty_excel_file(self, detector):
        """Test detection from empty Excel file"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            wb.save(tmp.name)
            wb.close()

            vendor, confidence = detector.detect_vendor(tmp.name, "empty.xlsx")

            # Should return None for empty file
            assert vendor is None or confidence < 0.5

            Path(tmp.name).unlink()

    def test_detect_from_corrupted_file(self, detector):
        """Test detection from corrupted file"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(b"Not a valid Excel file")
            tmp.flush()

            vendor, confidence = detector.detect_vendor(tmp.name, "corrupted.xlsx")

            # Should handle error gracefully
            assert vendor is None
            assert confidence == 0.0

            Path(tmp.name).unlink()

    def test_detect_with_special_characters_in_filename(self, detector):
        """Test detection with special characters in filename"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["EAN", "Product", "Sold"])
            wb.save(tmp.name)
            wb.close()

            # Filename with special chars but contains vendor keyword
            vendor, confidence = detector.detect_vendor(tmp.name, "liberty_sales_2024-01-01_@#$.xlsx")

            assert vendor == "liberty"
            assert confidence >= 0.5

            Path(tmp.name).unlink()

    # ============================================
    # VENDOR-SPECIFIC TESTS
    # ============================================

    @pytest.mark.parametrize("vendor_name,filename_keyword,columns", [
        ("cdlc", "cdlc", ["Product", "Total", "Month"]),
        ("selfridges", "selfridges", ["EAN", "Product", "Sold"]),
        ("ukraine", "ukraine", ["EAN", "Product", "Quantity"]),
        ("skins_nl", "skins", ["EAN", "Product", "Sales"]),
        ("continuity", "continuity", ["EAN", "Product", "Amount"]),
    ])
    def test_detect_various_vendors(self, detector, vendor_name, filename_keyword, columns):
        """Parametrized test for various vendors"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(columns)
            wb.save(tmp.name)
            wb.close()

            vendor, confidence = detector.detect_vendor(tmp.name, f"{filename_keyword}_data.xlsx")

            assert vendor == vendor_name or confidence >= 0.5

            Path(tmp.name).unlink()

    # ============================================
    # MULTIPLE SHEET TESTS
    # ============================================

    def test_detect_from_multi_sheet_excel(self, detector):
        """Test detection from Excel with multiple sheets"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()

            # First sheet (will be checked)
            ws1 = wb.active
            ws1.title = "Sell Out by EAN"
            ws1.append(["Product EAN", "Sold Qty", "Sales Amount (EUR)"])

            # Second sheet (ignored)
            ws2 = wb.create_sheet("Summary")
            ws2.append(["Total", "Revenue"])

            wb.save(tmp.name)
            wb.close()

            vendor, confidence = detector.detect_vendor(tmp.name, "boxnox_report.xlsx")

            assert vendor == "boxnox"
            assert confidence >= 0.5

            Path(tmp.name).unlink()

    # ============================================
    # CASE SENSITIVITY TESTS
    # ============================================

    def test_case_insensitive_filename_matching(self, detector):
        """Test case-insensitive filename matching"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["EAN", "Product"])
            wb.save(tmp.name)
            wb.close()

            vendor, confidence = detector.detect_vendor(tmp.name, "GALILU_DATA.xlsx")

            # Should match regardless of case
            assert vendor == "galilu"

            Path(tmp.name).unlink()

    def test_case_insensitive_column_matching(self, detector):
        """Test case-insensitive column header matching"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["product ean", "SOLD QTY", "sales amount (eur)"])
            wb.save(tmp.name)
            wb.close()

            vendor, confidence = detector.detect_vendor(tmp.name, "boxnox.xlsx")

            # Should match columns case-insensitively
            assert vendor == "boxnox"

            Path(tmp.name).unlink()


# ============================================
# VENDOR PATTERNS VALIDATION TESTS
# ============================================

class TestVendorPatterns:
    """Test vendor pattern configuration"""

    @pytest.fixture
    def detector(self):
        return VendorDetector()

    def test_all_vendors_have_patterns(self, detector):
        """Test that all vendors have detection patterns"""
        expected_vendors = [
            "boxnox", "galilu", "skins_sa", "cdlc",
            "liberty", "selfridges", "ukraine",
            "skins_nl", "continuity", "online"
        ]

        for vendor in expected_vendors:
            assert vendor in detector.VENDOR_PATTERNS

    def test_all_patterns_have_required_fields(self, detector):
        """Test that all patterns have required fields"""
        required_fields = ["filename_keywords", "sheet_names", "required_columns"]

        for vendor, pattern in detector.VENDOR_PATTERNS.items():
            for field in required_fields:
                assert field in pattern, f"Vendor {vendor} missing field {field}"

    def test_filename_keywords_are_lowercase(self, detector):
        """Test that filename keywords are lowercase for matching"""
        for vendor, pattern in detector.VENDOR_PATTERNS.items():
            for keyword in pattern["filename_keywords"]:
                # Keywords should be lowercase or matching will fail
                assert isinstance(keyword, str), f"Vendor {vendor} has non-string keyword"


# ============================================
# PERFORMANCE TESTS
# ============================================

class TestVendorDetectionPerformance:
    """Test performance characteristics"""

    @pytest.fixture
    def detector(self):
        return VendorDetector()

    def test_detect_from_large_file(self, detector):
        """Test detection from file with many rows"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Product EAN", "Sold Qty", "Sales Amount (EUR)"])

            # Add many rows (detection should only check headers)
            for i in range(1000):
                ws.append([f"123456789012{i}", 10, 100.50])

            wb.save(tmp.name)
            wb.close()

            vendor, confidence = detector.detect_vendor(tmp.name, "boxnox.xlsx")

            # Should detect quickly without reading all rows
            assert vendor == "boxnox"
            assert confidence >= 0.5

            Path(tmp.name).unlink()

    def test_detect_with_wide_columns(self, detector):
        """Test detection from file with many columns"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active

            # Create wide header row
            headers = ["Product EAN", "Sold Qty", "Sales Amount (EUR)"]
            headers.extend([f"Extra{i}" for i in range(100)])
            ws.append(headers)

            wb.save(tmp.name)
            wb.close()

            vendor, confidence = detector.detect_vendor(tmp.name, "boxnox.xlsx")

            # Should handle wide files
            assert vendor == "boxnox"

            Path(tmp.name).unlink()

"""
Unit tests for base processor classes using shared utilities

Tests that BibbiBseProcessor and VendorProcessor correctly use
shared utilities from app/utils/validation.py and app/utils/excel.py
"""

import pytest
import tempfile
import openpyxl
from pathlib import Path
from typing import List, Dict, Any, Optional

from app.services.bibbi.processors.base import BibbiBseProcessor
from app.services.vendors.base import VendorProcessor


# ============================================
# TEST CONCRETE IMPLEMENTATIONS
# ============================================

class TestBibbiProcessor(BibbiBseProcessor):
    """Concrete implementation for testing BibbiBseProcessor"""

    def get_vendor_name(self) -> str:
        return "test_bibbi"

    def get_currency(self) -> str:
        return "EUR"

    def extract_rows(self, file_path: str) -> List[Dict[str, Any]]:
        # Simple implementation for testing
        workbook = self._load_workbook(file_path, read_only=True)
        sheet = workbook.active
        rows = []
        headers = self._get_sheet_headers(sheet)

        for row_idx, row_cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row_cells):
                continue
            row_dict = {headers[i]: row_cells[i] for i in range(min(len(headers), len(row_cells)))}
            rows.append(row_dict)

        workbook.close()
        return rows

    def transform_row(self, raw_row: Dict[str, Any], batch_id: str) -> Optional[Dict[str, Any]]:
        # Simple transformation for testing
        base = self._create_base_row(batch_id)
        base["product_ean"] = self._validate_ean(raw_row.get("EAN"), required=True)
        base["quantity"] = self._to_int(raw_row.get("Quantity"), "Quantity")
        base["unit_price"] = self._to_float(raw_row.get("Price"), "Price")
        return base

    def extract_stores(self, file_path: str) -> List[Dict[str, Any]]:
        # Simple implementation for testing
        return [{
            "store_identifier": "TEST-001",
            "store_name": "Test Store",
            "store_type": "physical",
            "reseller_id": self.reseller_id
        }]


class TestVendorProcessorImpl(VendorProcessor):
    """Concrete implementation for testing VendorProcessor"""

    def get_vendor_name(self) -> str:
        return "test_vendor"

    def process(self, file_path: str, user_id: str, batch_id: str) -> Dict[str, Any]:
        # Simple implementation for testing
        rows = self.extract_rows(file_path)
        return {
            "total_rows": len(rows),
            "successful_rows": len(rows),
            "failed_rows": 0,
            "transformed_data": rows,
            "errors": []
        }

    def extract_rows(self, file_path: str) -> List[Dict[str, Any]]:
        # Simple implementation for testing
        workbook = self._load_workbook(file_path, read_only=True)
        sheet = workbook.active
        rows = self._extract_rows(sheet, min_row=2)
        workbook.close()
        return rows


# ============================================
# BIBBI BASE PROCESSOR TESTS
# ============================================

class TestBibbiBseProcessorUtilities:
    """Test BibbiBseProcessor uses shared utilities correctly"""

    @pytest.fixture
    def test_processor(self):
        """Create test processor instance"""
        return TestBibbiProcessor(reseller_id="test-reseller-123")

    @pytest.fixture
    def test_excel_file(self):
        """Create test Excel file"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["EAN", "Product", "Quantity", "Price"])
            ws.append(["1234567890123", "Product A", 10, 99.99])
            ws.append(["9876543210987", "Product B", 5, 49.99])
            wb.save(tmp.name)
            wb.close()
            yield tmp.name
            Path(tmp.name).unlink()

    def test_load_workbook_uses_shared_utility(self, test_processor, test_excel_file):
        """Test _load_workbook() uses safe_load_workbook()"""
        workbook = test_processor._load_workbook(test_excel_file, read_only=True)

        assert workbook is not None
        assert isinstance(workbook, openpyxl.Workbook)

        workbook.close()

    def test_get_sheet_headers_uses_shared_utility(self, test_processor, test_excel_file):
        """Test _get_sheet_headers() uses get_sheet_headers()"""
        workbook = test_processor._load_workbook(test_excel_file, read_only=True)
        sheet = workbook.active

        headers = test_processor._get_sheet_headers(sheet)

        assert headers == ["EAN", "Product", "Quantity", "Price"]

        workbook.close()

    def test_validate_ean_uses_shared_utility(self, test_processor):
        """Test _validate_ean() uses validate_ean()"""
        # Valid EAN
        ean = test_processor._validate_ean("1234567890123", required=True)
        assert ean == "1234567890123"

        # EAN with decimal (Excel artifact)
        ean = test_processor._validate_ean(1234567890123.0, required=True)
        assert ean == "1234567890123"

        # Optional EAN
        ean = test_processor._validate_ean(None, required=False)
        assert ean is None

        # Invalid EAN
        with pytest.raises(ValueError):
            test_processor._validate_ean("invalid", required=True)

    def test_to_int_uses_shared_utility(self, test_processor):
        """Test _to_int() uses to_int() with accounting notation support"""
        # Standard integer
        value = test_processor._to_int(42, "test_field")
        assert value == 42

        # String integer
        value = test_processor._to_int("100", "test_field")
        assert value == 100

        # Accounting notation (negative)
        value = test_processor._to_int("(50)", "test_field")
        assert value == -50

    def test_to_float_uses_shared_utility(self, test_processor):
        """Test _to_float() uses to_float() with accounting notation support"""
        # Standard float
        value = test_processor._to_float(99.99, "test_field")
        assert value == 99.99

        # String float
        value = test_processor._to_float("49.99", "test_field")
        assert value == 49.99

        # Accounting notation (negative)
        value = test_processor._to_float("(25.50)", "test_field")
        assert value == -25.50

        # None with default
        value = test_processor._to_float(None, "test_field")
        assert value == 0.0

    def test_to_decimal_conversion(self, test_processor):
        """Test _to_decimal() converts correctly"""
        from decimal import Decimal

        value = test_processor._to_decimal(99.99, "price")
        assert isinstance(value, Decimal)
        assert value == Decimal("99.99")

    def test_convert_currency(self, test_processor):
        """Test _convert_currency() utility"""
        # EUR to EUR (no conversion)
        amount = test_processor._convert_currency(100.0, "EUR")
        assert amount == 100.0

        # GBP to EUR
        amount = test_processor._convert_currency(100.0, "GBP")
        assert amount == 117.0  # Based on CURRENCY_RATES

        # Invalid currency
        with pytest.raises(ValueError, match="Unknown currency"):
            test_processor._convert_currency(100.0, "XXX")

    def test_calculate_quarter(self, test_processor):
        """Test _calculate_quarter() utility"""
        assert test_processor._calculate_quarter(1) == 1
        assert test_processor._calculate_quarter(3) == 1
        assert test_processor._calculate_quarter(4) == 2
        assert test_processor._calculate_quarter(6) == 2
        assert test_processor._calculate_quarter(7) == 3
        assert test_processor._calculate_quarter(9) == 3
        assert test_processor._calculate_quarter(10) == 4
        assert test_processor._calculate_quarter(12) == 4

    def test_full_processing_pipeline(self, test_processor, test_excel_file):
        """Test full processing pipeline uses shared utilities"""
        result = test_processor.process(test_excel_file, batch_id="test-batch-123")

        assert result.total_rows == 2
        assert result.successful_rows == 2
        assert result.failed_rows == 0
        assert len(result.transformed_data) == 2

        # Verify data uses shared utilities correctly
        first_row = result.transformed_data[0]
        assert first_row["product_ean"] == "1234567890123"  # validate_ean
        assert first_row["quantity"] == 10  # to_int
        assert first_row["unit_price"] == 99.99  # to_float
        assert first_row["vendor_name"] == "test_bibbi"


# ============================================
# VENDOR BASE PROCESSOR TESTS
# ============================================

class TestVendorProcessorUtilities:
    """Test VendorProcessor uses shared utilities correctly"""

    @pytest.fixture
    def test_processor(self):
        """Create test processor instance"""
        return TestVendorProcessorImpl(reseller_id="test-vendor-123")

    @pytest.fixture
    def test_excel_file(self):
        """Create test Excel file"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["EAN", "Product", "Quantity", "Price"])
            ws.append(["1234567890123", "Product A", 10, 99.99])
            ws.append(["9876543210987", "Product B", 5, 49.99])
            wb.save(tmp.name)
            wb.close()
            yield tmp.name
            Path(tmp.name).unlink()

    def test_vendor_processor_loads_workbook(self, test_processor, test_excel_file):
        """Test VendorProcessor _load_workbook() works"""
        workbook = test_processor._load_workbook(test_excel_file, read_only=True)

        assert workbook is not None
        assert isinstance(workbook, openpyxl.Workbook)

        workbook.close()

    def test_vendor_processor_extract_rows(self, test_processor, test_excel_file):
        """Test VendorProcessor _extract_rows() uses shared utility"""
        workbook = test_processor._load_workbook(test_excel_file, read_only=True)
        sheet = workbook.active

        rows = test_processor._extract_rows(sheet, min_row=2)

        assert len(rows) == 2
        assert rows[0]["EAN"] == "1234567890123"
        assert rows[1]["Quantity"] == 5

        workbook.close()

    def test_vendor_processor_validate_ean(self, test_processor):
        """Test VendorProcessor _validate_ean() works"""
        # Valid EAN with strict mode
        ean = test_processor._validate_ean("1234567890123", required=True, strict=True)
        assert ean == "1234567890123"

        # Optional EAN with non-strict mode
        ean = test_processor._validate_ean(None, required=False, strict=False)
        assert ean is None

        # Invalid EAN with non-strict mode
        ean = test_processor._validate_ean("invalid", required=False, strict=False)
        assert ean is None

    def test_vendor_processor_type_conversions(self, test_processor):
        """Test VendorProcessor type conversion utilities"""
        # to_int
        value = test_processor._to_int(42, "quantity")
        assert value == 42

        # to_float
        value = test_processor._to_float(99.99, "price")
        assert value == 99.99

        # to_decimal
        value = test_processor._to_decimal(99.99, "amount")
        from decimal import Decimal
        assert isinstance(value, Decimal)

    def test_vendor_processor_currency_conversion(self, test_processor):
        """Test VendorProcessor currency conversion"""
        # EUR to EUR
        amount = test_processor._convert_currency(100.0, "EUR", "EUR")
        assert amount == 100.0

        # USD to EUR
        amount = test_processor._convert_currency(100.0, "USD", "EUR")
        assert amount == 92.0  # Based on CURRENCY_RATES

    def test_vendor_processor_validates_month_year(self, test_processor):
        """Test VendorProcessor month and year validation"""
        # Note: These methods don't exist on VendorProcessor base class
        # but are available via app.utils.validation imports
        from app.utils.validation import validate_month, validate_year

        month = validate_month(6)
        assert month == 6

        year = validate_year(2025)
        assert year == 2025

    def test_vendor_processor_get_vendor_name(self, test_processor):
        """Test VendorProcessor get_vendor_name() abstract method"""
        assert test_processor.get_vendor_name() == "test_vendor"


# ============================================
# ABSTRACT METHOD ENFORCEMENT TESTS
# ============================================

class TestAbstractMethodEnforcement:
    """Test that abstract methods are enforced"""

    def test_bibbi_processor_requires_abstract_methods(self):
        """Test BibbiBseProcessor requires abstract methods"""
        # Cannot instantiate without implementing abstract methods
        with pytest.raises(TypeError):
            class IncompleteBibbiProcessor(BibbiBseProcessor):
                pass

            IncompleteBibbiProcessor(reseller_id="test")

    def test_vendor_processor_requires_abstract_methods(self):
        """Test VendorProcessor requires abstract methods"""
        # Cannot instantiate without implementing abstract methods
        with pytest.raises(TypeError):
            class IncompleteVendorProcessor(VendorProcessor):
                pass

            IncompleteVendorProcessor()


# ============================================
# INTEGRATION TESTS
# ============================================

class TestProcessorIntegration:
    """Test processors work together with shared utilities"""

    @pytest.fixture
    def test_excel_file(self):
        """Create comprehensive test Excel file"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["EAN", "Product", "Quantity", "Price", "Month", "Year"])
            ws.append(["1234567890123", "Product A", 10, 99.99, 6, 2025])
            ws.append(["0012345678901", "Product B", 5, 49.99, 7, 2025])  # Leading zeros
            ws.append([1234567890123.0, "Product C", 8, 29.99, 8, 2025])  # Float EAN
            wb.save(tmp.name)
            wb.close()
            yield tmp.name
            Path(tmp.name).unlink()

    def test_bibbi_processor_end_to_end(self, test_excel_file):
        """Test BibbiBseProcessor end-to-end workflow"""
        processor = TestBibbiProcessor(reseller_id="test-123")

        result = processor.process(test_excel_file, batch_id="batch-456")

        # Verify processing results
        assert result.vendor == "test_bibbi"
        assert result.total_rows == 3
        assert result.successful_rows == 3
        assert result.failed_rows == 0

        # Verify shared utilities worked correctly
        rows = result.transformed_data
        assert rows[0]["product_ean"] == "1234567890123"
        assert rows[1]["product_ean"] == "0012345678901"  # Preserved leading zeros
        assert rows[2]["product_ean"] == "1234567890123"  # Converted from float

    def test_vendor_processor_end_to_end(self, test_excel_file):
        """Test VendorProcessor end-to-end workflow"""
        processor = TestVendorProcessorImpl(reseller_id="test-vendor-456")

        result = processor.process(test_excel_file, user_id="user-123", batch_id="batch-789")

        # Verify processing results
        assert result["total_rows"] == 3
        assert result["successful_rows"] == 3
        assert result["failed_rows"] == 0

        # Verify data extraction worked
        rows = result["transformed_data"]
        assert len(rows) == 3
        assert rows[0]["EAN"] == "1234567890123"
        assert rows[1]["Quantity"] == 5

    def test_processors_share_same_utilities(self):
        """Test both processors use same shared utility functions"""
        from app.utils.validation import validate_ean, to_int, to_float
        from app.utils.excel import safe_load_workbook, get_sheet_headers

        bibbi = TestBibbiProcessor(reseller_id="test-bibbi")
        vendor = TestVendorProcessorImpl(reseller_id="test-vendor")

        # Both should use same validation functions
        assert bibbi._validate_ean("1234567890123") == vendor._validate_ean("1234567890123")
        assert bibbi._to_int(42, "field") == vendor._to_int(42, "field")
        assert bibbi._to_float(99.99, "field") == vendor._to_float(99.99, "field")

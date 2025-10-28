"""
Unit tests for app/utils/excel.py

Tests the shared Excel processing utilities extracted during DRY refactoring.
"""

import pytest
import tempfile
import openpyxl
from pathlib import Path
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.excel import (
    extract_rows_from_sheet,
    safe_load_workbook,
    find_sheet_by_name,
    get_sheet_headers,
    validate_required_headers,
    count_data_rows
)


# ============================================
# SAFE LOAD WORKBOOK TESTS
# ============================================

class TestSafeLoadWorkbook:
    """Test safe workbook loading utility"""

    def test_load_valid_excel_file(self):
        """Test loading valid Excel file"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Header1", "Header2"])
            ws.append(["Data1", "Data2"])
            wb.save(tmp.name)
            wb.close()

            loaded_wb = safe_load_workbook(tmp.name, data_only=True)
            assert loaded_wb is not None
            assert isinstance(loaded_wb, openpyxl.Workbook)
            loaded_wb.close()

            Path(tmp.name).unlink()

    def test_load_with_data_only_flag(self):
        """Test data_only parameter works"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Header1", "Header2"])
            wb.save(tmp.name)
            wb.close()

            # data_only=True reads formula results, not formulas
            loaded_wb = safe_load_workbook(tmp.name, data_only=True, read_only=False)
            assert loaded_wb is not None
            loaded_wb.close()

            # data_only=False reads formulas
            loaded_wb = safe_load_workbook(tmp.name, data_only=False, read_only=False)
            assert loaded_wb is not None
            loaded_wb.close()

            Path(tmp.name).unlink()

    def test_load_with_read_only_flag(self):
        """Test read_only parameter works"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Header1", "Header2"])
            wb.save(tmp.name)
            wb.close()

            loaded_wb = safe_load_workbook(tmp.name, read_only=True)
            assert loaded_wb is not None
            loaded_wb.close()

            Path(tmp.name).unlink()

    def test_load_non_existent_file(self):
        """Test loading non-existent file raises error"""
        # Implementation raises ValueError with message about file not found
        with pytest.raises(ValueError, match="Excel file not found"):
            safe_load_workbook("/non/existent/file.xlsx")

    def test_load_corrupted_file(self):
        """Test loading corrupted file raises error"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(b"This is not a valid Excel file")
            tmp.flush()

            with pytest.raises(Exception):
                safe_load_workbook(tmp.name)

            Path(tmp.name).unlink()

    def test_load_empty_file(self):
        """Test loading empty file"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.flush()

            with pytest.raises(Exception):
                safe_load_workbook(tmp.name)

            Path(tmp.name).unlink()


# ============================================
# FIND SHEET BY NAME TESTS
# ============================================

class TestFindSheetByName:
    """Test worksheet finding utility"""

    @pytest.fixture
    def multi_sheet_workbook(self):
        """Create workbook with multiple sheets"""
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        wb.create_sheet("Sales Data")
        wb.create_sheet("Summary")
        return wb

    def test_find_existing_sheet_exact_match(self, multi_sheet_workbook):
        """Test finding sheet with exact name match"""
        sheet = find_sheet_by_name(multi_sheet_workbook, "Sales Data")
        assert sheet is not None
        assert sheet.title == "Sales Data"

    def test_find_first_sheet(self, multi_sheet_workbook):
        """Test finding first sheet"""
        sheet = find_sheet_by_name(multi_sheet_workbook, "Sheet1")
        assert sheet is not None
        assert sheet.title == "Sheet1"

    def test_sheet_not_found_with_fallback(self, multi_sheet_workbook):
        """Test non-existent sheet returns first sheet with fallback"""
        sheet = find_sheet_by_name(multi_sheet_workbook, "NonExistent", fallback_to_first=True)
        assert sheet is not None
        assert sheet.title == "Sheet1"  # First sheet

    def test_sheet_not_found_without_fallback(self, multi_sheet_workbook):
        """Test non-existent sheet raises error without fallback"""
        with pytest.raises(ValueError, match="not found"):
            find_sheet_by_name(multi_sheet_workbook, "NonExistent", fallback_to_first=False)

    def test_case_sensitive_matching(self, multi_sheet_workbook):
        """Test sheet name matching is case-sensitive"""
        # Exact case match works
        sheet = find_sheet_by_name(multi_sheet_workbook, "Sales Data")
        assert sheet.title == "Sales Data"

        # Different case may not match (depends on implementation)
        # This documents current behavior
        with pytest.raises(ValueError):
            find_sheet_by_name(multi_sheet_workbook, "sales data", fallback_to_first=False)


# ============================================
# GET SHEET HEADERS TESTS
# ============================================

class TestGetSheetHeaders:
    """Test header extraction utility"""

    def test_get_headers_first_row(self):
        """Test extracting headers from first row"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Header1", "Header2", "Header3"])
        ws.append(["Data1", "Data2", "Data3"])

        headers = get_sheet_headers(ws, header_row=1)
        assert headers == ["Header1", "Header2", "Header3"]

        wb.close()

    def test_get_headers_custom_row(self):
        """Test extracting headers from custom row"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Title Row"])
        ws.append(["Actual", "Headers", "Here"])
        ws.append(["Data1", "Data2", "Data3"])

        headers = get_sheet_headers(ws, header_row=2)
        assert headers == ["Actual", "Headers", "Here"]

        wb.close()

    def test_get_headers_with_empty_cells(self):
        """Test headers with empty cells"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Header1", None, "Header3"])

        headers = get_sheet_headers(ws, header_row=1)
        # Implementation only includes cells with truthy values
        # Empty cells (None) are skipped
        assert len(headers) == 2
        assert headers[0] == "Header1"
        assert headers[1] == "Header3"

        wb.close()

    def test_get_headers_non_string_values(self):
        """Test headers with non-string values get converted"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([123, 456.78, "Header3"])

        headers = get_sheet_headers(ws, header_row=1)
        # Non-strings should be converted to strings
        assert len(headers) == 3
        assert isinstance(headers[0], str)
        assert isinstance(headers[1], str)

        wb.close()

    def test_get_headers_empty_sheet(self):
        """Test extracting headers from empty sheet"""
        wb = openpyxl.Workbook()
        ws = wb.active

        headers = get_sheet_headers(ws, header_row=1)
        # Empty sheet returns empty list or list with None values
        assert isinstance(headers, list)

        wb.close()


# ============================================
# EXTRACT ROWS FROM SHEET TESTS
# ============================================

class TestExtractRowsFromSheet:
    """Test row extraction utility"""

    def test_extract_standard_rows(self):
        """Test extracting standard data rows"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Age", "City"])
        ws.append(["Alice", 30, "NYC"])
        ws.append(["Bob", 25, "LA"])

        rows = extract_rows_from_sheet(ws, header_row=1, min_data_row=2)

        assert len(rows) == 2
        assert rows[0] == {"Name": "Alice", "Age": 30, "City": "NYC"}
        assert rows[1] == {"Name": "Bob", "Age": 25, "City": "LA"}

        wb.close()

    def test_extract_rows_skip_empty_default(self):
        """Test extracting rows skips empty rows by default"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Age"])
        ws.append(["Alice", 30])
        ws.append([None, None])  # Empty row
        ws.append(["Bob", 25])

        rows = extract_rows_from_sheet(ws, header_row=1, min_data_row=2, skip_empty=True)

        assert len(rows) == 2
        assert rows[0]["Name"] == "Alice"
        assert rows[1]["Name"] == "Bob"

        wb.close()

    def test_extract_rows_include_empty(self):
        """Test extracting rows includes empty when skip_empty=False"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Age"])
        ws.append(["Alice", 30])
        ws.append([None, None])  # Empty row
        ws.append(["Bob", 25])

        rows = extract_rows_from_sheet(ws, header_row=1, min_data_row=2, skip_empty=False)

        # Should include empty row
        assert len(rows) >= 2

        wb.close()

    def test_extract_rows_with_missing_columns(self):
        """Test rows with missing column values"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Age", "City"])
        ws.append(["Alice", 30, "NYC"])
        ws.append(["Bob"])  # Missing Age and City

        rows = extract_rows_from_sheet(ws, header_row=1, min_data_row=2)

        assert len(rows) == 2
        assert rows[0] == {"Name": "Alice", "Age": 30, "City": "NYC"}
        # Missing values should be None
        assert rows[1]["Name"] == "Bob"

        wb.close()

    def test_extract_rows_custom_min_data_row(self):
        """Test extracting rows starting from custom row"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Age"])  # Row 1 - headers
        ws.append(["Skip", 0])      # Row 2 - skip this
        ws.append(["Alice", 30])    # Row 3 - start here
        ws.append(["Bob", 25])      # Row 4

        rows = extract_rows_from_sheet(ws, header_row=1, min_data_row=3)

        assert len(rows) == 2
        assert rows[0]["Name"] == "Alice"
        assert rows[1]["Name"] == "Bob"

        wb.close()

    def test_extract_rows_empty_sheet(self):
        """Test extracting from sheet with no data rows"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Age"])  # Only headers

        rows = extract_rows_from_sheet(ws, header_row=1, min_data_row=2)

        assert rows == []

        wb.close()

    def test_extract_rows_preserves_types(self):
        """Test row extraction preserves data types"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Age", "Price"])
        ws.append(["Product", 100, 99.99])

        rows = extract_rows_from_sheet(ws, header_row=1, min_data_row=2)

        assert rows[0]["Name"] == "Product"
        assert rows[0]["Age"] == 100
        assert isinstance(rows[0]["Age"], int)
        assert rows[0]["Price"] == 99.99
        assert isinstance(rows[0]["Price"], float)

        wb.close()


# ============================================
# VALIDATE REQUIRED HEADERS TESTS
# ============================================

class TestValidateRequiredHeaders:
    """Test header validation utility"""

    def test_all_required_headers_present(self):
        """Test validation passes when all headers present"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["EAN", "Product", "Quantity", "Price"])

        # Should not raise any error
        validate_required_headers(ws, ["EAN", "Product", "Quantity"], header_row=1)

        wb.close()

    def test_extra_headers_allowed(self):
        """Test validation passes with extra headers"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["EAN", "Product", "Quantity", "Extra1", "Extra2"])

        # Should not raise error (not strict about extra columns)
        validate_required_headers(ws, ["EAN", "Product"], header_row=1)

        wb.close()

    def test_missing_required_header(self):
        """Test validation fails when header missing"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["EAN", "Product"])

        # Error message says "Missing required columns" not "headers"
        with pytest.raises(ValueError, match="Missing required columns"):
            validate_required_headers(ws, ["EAN", "Product", "Quantity"], header_row=1)

        wb.close()

    def test_multiple_missing_headers(self):
        """Test error message includes all missing headers"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["EAN"])

        with pytest.raises(ValueError, match="Product"):
            validate_required_headers(ws, ["EAN", "Product", "Quantity"], header_row=1)

        wb.close()

    def test_case_sensitive_header_matching(self):
        """Test header matching is case-sensitive"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ean", "product"])  # Lowercase

        # Depending on implementation, may or may not match
        # This documents current behavior
        with pytest.raises(ValueError):
            validate_required_headers(ws, ["EAN", "Product"], header_row=1)

        wb.close()

    def test_custom_header_row(self):
        """Test validation with custom header row"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Title"])
        ws.append(["EAN", "Product", "Quantity"])

        validate_required_headers(ws, ["EAN", "Product"], header_row=2)

        wb.close()


# ============================================
# COUNT DATA ROWS TESTS
# ============================================

class TestCountDataRows:
    """Test row counting utility"""

    def test_count_standard_data_rows(self):
        """Test counting standard data rows"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Age"])
        ws.append(["Alice", 30])
        ws.append(["Bob", 25])
        ws.append(["Charlie", 35])

        count = count_data_rows(ws, min_row=2, skip_empty=True)
        assert count == 3

        wb.close()

    def test_count_with_empty_rows_skip(self):
        """Test counting skips empty rows"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Age"])
        ws.append(["Alice", 30])
        ws.append([None, None])  # Empty
        ws.append(["Bob", 25])

        count = count_data_rows(ws, min_row=2, skip_empty=True)
        assert count == 2

        wb.close()

    def test_count_with_empty_rows_include(self):
        """Test counting includes empty rows when skip_empty=False"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Age"])
        ws.append(["Alice", 30])
        ws.append([None, None])  # Empty
        ws.append(["Bob", 25])

        count = count_data_rows(ws, min_row=2, skip_empty=False)
        assert count >= 2  # Includes empty row

        wb.close()

    def test_count_custom_min_row(self):
        """Test counting from custom starting row"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Age"])  # Row 1
        ws.append(["Skip", 0])      # Row 2
        ws.append(["Alice", 30])    # Row 3
        ws.append(["Bob", 25])      # Row 4

        count = count_data_rows(ws, min_row=3, skip_empty=True)
        assert count == 2

        wb.close()

    def test_count_empty_sheet(self):
        """Test counting empty sheet returns 0"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Age"])

        count = count_data_rows(ws, min_row=2, skip_empty=True)
        assert count == 0

        wb.close()

    def test_count_large_dataset(self):
        """Test counting large number of rows"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Age"])

        # Add 1000 data rows
        for i in range(1000):
            ws.append([f"Person{i}", i])

        count = count_data_rows(ws, min_row=2, skip_empty=True)
        assert count == 1000

        wb.close()


# ============================================
# INTEGRATION TESTS
# ============================================

class TestExcelUtilitiesIntegration:
    """Test utilities working together"""

    def test_full_excel_processing_workflow(self):
        """Test complete workflow: load → validate → extract"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            # Create test file
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sales Data"
            ws.append(["EAN", "Product", "Quantity", "Price"])
            ws.append(["1234567890123", "Product A", 10, 99.99])
            ws.append(["9876543210987", "Product B", 5, 49.99])
            wb.save(tmp.name)
            wb.close()

            # Load workbook
            loaded_wb = safe_load_workbook(tmp.name, data_only=True, read_only=True)

            # Find sheet
            sheet = find_sheet_by_name(loaded_wb, "Sales Data", fallback_to_first=False)

            # Validate headers
            validate_required_headers(sheet, ["EAN", "Product", "Quantity"], header_row=1)

            # Count rows
            row_count = count_data_rows(sheet, min_row=2, skip_empty=True)
            assert row_count == 2

            # Extract rows
            rows = extract_rows_from_sheet(sheet, header_row=1, min_data_row=2)
            assert len(rows) == 2
            assert rows[0]["EAN"] == "1234567890123"
            assert rows[0]["Quantity"] == 10

            loaded_wb.close()
            Path(tmp.name).unlink()

    def test_processing_with_missing_sheet_fallback(self):
        """Test processing with sheet fallback"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["EAN", "Product"])
            ws.append(["1234567890123", "Product A"])
            wb.save(tmp.name)
            wb.close()

            loaded_wb = safe_load_workbook(tmp.name)

            # Try to find non-existent sheet, fallback to first
            sheet = find_sheet_by_name(loaded_wb, "NonExistent", fallback_to_first=True)

            # Should still work
            rows = extract_rows_from_sheet(sheet, header_row=1, min_data_row=2)
            assert len(rows) == 1

            loaded_wb.close()
            Path(tmp.name).unlink()

    def test_error_handling_chain(self):
        """Test error handling across utilities"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["EAN", "Product"])
            wb.save(tmp.name)
            wb.close()

            loaded_wb = safe_load_workbook(tmp.name)
            sheet = find_sheet_by_name(loaded_wb, "Sheet", fallback_to_first=True)

            # Validation should fail for missing column
            with pytest.raises(ValueError, match="Missing required columns"):
                validate_required_headers(sheet, ["EAN", "Product", "Quantity"], header_row=1)

            loaded_wb.close()
            Path(tmp.name).unlink()

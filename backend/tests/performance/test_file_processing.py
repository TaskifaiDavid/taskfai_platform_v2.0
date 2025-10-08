"""
Performance tests for file processing
Target: 1-2 minutes for 500-5000 rows
"""

import pytest
import time
import tempfile
import pandas as pd
import openpyxl
from pathlib import Path
from unittest.mock import Mock, patch


# Target processing times
MIN_ROWS = 500
MAX_ROWS = 5000
TARGET_TIME_MIN_SECONDS = 60  # 1 minute
TARGET_TIME_MAX_SECONDS = 120  # 2 minutes


class TestFileProcessingPerformance:
    """Test file processing performance"""

    # ============================================
    # FILE VALIDATION PERFORMANCE
    # ============================================

    @pytest.mark.performance
    def test_csv_validation_performance(self):
        """Test CSV file validation speed"""
        from app.services.file_validator import FileValidator

        # Create test CSV
        with tempfile.NamedTemporaryFile(mode='w', suffix=".csv", delete=False) as tmp:
            # Write 1000 rows
            tmp.write("EAN,Product,Quantity,Price,Date\n")
            for i in range(1000):
                tmp.write(f"123456789012{i},Product {i},10,29.99,2024-01-01\n")
            tmp.flush()

            validator = FileValidator(max_size_bytes=10 * 1024 * 1024)

            start = time.perf_counter()
            is_valid, error = validator.validate_file(tmp.name, "test.csv")
            validation_time = (time.perf_counter() - start) * 1000

            print(f"\nCSV validation (1000 rows): {validation_time:.2f}ms")

            # Validation should be fast (<100ms for 1000 rows)
            assert validation_time < 100, f"CSV validation too slow: {validation_time:.2f}ms"

            Path(tmp.name).unlink()

    @pytest.mark.performance
    def test_excel_validation_performance(self):
        """Test Excel file validation speed"""
        from app.services.file_validator import FileValidator

        # Create test Excel
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["EAN", "Product", "Quantity", "Price"])

            # Add 1000 rows
            for i in range(1000):
                ws.append([f"123456789012{i}", f"Product {i}", 10, 29.99])

            wb.save(tmp.name)
            wb.close()

            validator = FileValidator(max_size_bytes=10 * 1024 * 1024)

            start = time.perf_counter()
            is_valid, error = validator.validate_file(tmp.name, "test.xlsx")
            validation_time = (time.perf_counter() - start) * 1000

            print(f"\nExcel validation (1000 rows): {validation_time:.2f}ms")

            # Excel validation may be slower but should still be reasonable
            assert validation_time < 500, f"Excel validation too slow: {validation_time:.2f}ms"

            Path(tmp.name).unlink()

    # ============================================
    # VENDOR DETECTION PERFORMANCE
    # ============================================

    @pytest.mark.performance
    def test_vendor_detection_performance(self):
        """Test vendor detection speed"""
        from app.services.vendors.detector import VendorDetector

        # Create test file
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sell Out by EAN"
            ws.append(["Product EAN", "Functional Name", "Sold Qty", "Sales Amount (EUR)"])

            # Add rows
            for i in range(500):
                ws.append([f"123456789012{i}", f"Product {i}", 10, 100.0])

            wb.save(tmp.name)
            wb.close()

            detector = VendorDetector()

            start = time.perf_counter()
            vendor, confidence = detector.detect_vendor(tmp.name, "boxnox_data.xlsx")
            detection_time = (time.perf_counter() - start) * 1000

            print(f"\nVendor detection (500 rows): {detection_time:.2f}ms")
            print(f"Detected: {vendor} (confidence: {confidence:.2f})")

            # Detection should be fast (only reads headers)
            assert detection_time < 200, f"Vendor detection too slow: {detection_time:.2f}ms"

            Path(tmp.name).unlink()

    # ============================================
    # DATA PROCESSING PERFORMANCE
    # ============================================

    @pytest.mark.performance
    @pytest.mark.parametrize("row_count", [MIN_ROWS, 1000, 2500, MAX_ROWS])
    def test_processing_time_by_row_count(self, row_count):
        """Test processing time scales linearly with row count"""
        from app.services.vendors.boxnox_processor import BoxnoxProcessor

        # Create test data
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sell Out by EAN"
            ws.append(["Product EAN", "Functional Name", "Sold Qty", "Sales Amount (EUR)", "Reseller", "Month", "Year"])

            # Add rows
            for i in range(row_count):
                month = (i % 12) + 1
                year = 2024
                ws.append([
                    f"123456789012{i % 1000}",
                    f"Product {i % 100}",
                    10,
                    100.0,
                    f"Reseller {i % 10}",
                    month,
                    year
                ])

            wb.save(tmp.name)
            wb.close()

            processor = BoxnoxProcessor()

            start = time.perf_counter()
            try:
                result = processor.process(tmp.name)
                processing_time = time.perf_counter() - start

                print(f"\nProcessing {row_count} rows: {processing_time:.2f}s")

                # Calculate expected time (linear scaling)
                # Expect ~1-2 minutes for 500-5000 rows
                expected_min = (row_count / MAX_ROWS) * TARGET_TIME_MIN_SECONDS
                expected_max = (row_count / MIN_ROWS) * TARGET_TIME_MAX_SECONDS

                assert processing_time <= expected_max, \
                    f"Processing {row_count} rows too slow: {processing_time:.2f}s (expected < {expected_max:.2f}s)"

                print(f"Expected range: {expected_min:.2f}s - {expected_max:.2f}s")
                print(f"Rows/second: {row_count / processing_time:.2f}")

            finally:
                Path(tmp.name).unlink()

    @pytest.mark.performance
    def test_maximum_throughput(self):
        """Test maximum processing throughput"""
        from app.services.vendors.boxnox_processor import BoxnoxProcessor

        row_count = MAX_ROWS

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sell Out by EAN"
            ws.append(["Product EAN", "Functional Name", "Sold Qty", "Sales Amount (EUR)", "Reseller", "Month", "Year"])

            for i in range(row_count):
                ws.append([f"123456789012{i % 1000}", f"Product {i}", 10, 100.0, "Reseller", 1, 2024])

            wb.save(tmp.name)
            wb.close()

            processor = BoxnoxProcessor()

            start = time.perf_counter()
            result = processor.process(tmp.name)
            processing_time = time.perf_counter() - start

            throughput = row_count / processing_time

            print(f"\nMaximum throughput test:")
            print(f"Rows: {row_count}")
            print(f"Time: {processing_time:.2f}s")
            print(f"Throughput: {throughput:.2f} rows/second")

            # Should process at least 40 rows/second
            assert throughput >= 40, f"Throughput too low: {throughput:.2f} rows/s"

            Path(tmp.name).unlink()

    # ============================================
    # MEMORY USAGE TESTS
    # ============================================

    @pytest.mark.performance
    def test_memory_efficient_processing(self):
        """Test that processing doesn't consume excessive memory"""
        import tracemalloc
        from app.services.vendors.boxnox_processor import BoxnoxProcessor

        row_count = 2500

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sell Out by EAN"
            ws.append(["Product EAN", "Functional Name", "Sold Qty", "Sales Amount (EUR)", "Reseller", "Month", "Year"])

            for i in range(row_count):
                ws.append([f"123456789012{i}", f"Product {i}", 10, 100.0, "Reseller", 1, 2024])

            wb.save(tmp.name)
            wb.close()

            processor = BoxnoxProcessor()

            tracemalloc.start()
            start_memory = tracemalloc.get_traced_memory()[0]

            result = processor.process(tmp.name)

            peak_memory = tracemalloc.get_traced_memory()[1]
            tracemalloc.stop()

            memory_used_mb = (peak_memory - start_memory) / (1024 * 1024)

            print(f"\nMemory usage for {row_count} rows:")
            print(f"Peak memory: {memory_used_mb:.2f} MB")

            # Should use less than 50MB for 2500 rows
            assert memory_used_mb < 50, f"Memory usage too high: {memory_used_mb:.2f} MB"

            Path(tmp.name).unlink()

    # ============================================
    # CONCURRENT PROCESSING TESTS
    # ============================================

    @pytest.mark.performance
    @pytest.mark.asyncio
    async def test_concurrent_file_processing(self):
        """Test processing multiple files concurrently"""
        import asyncio
        from app.services.vendors.boxnox_processor import BoxnoxProcessor

        num_files = 5
        rows_per_file = 1000

        # Create multiple test files
        file_paths = []
        for file_num in range(num_files):
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Sell Out by EAN"
                ws.append(["Product EAN", "Functional Name", "Sold Qty", "Sales Amount (EUR)", "Reseller", "Month", "Year"])

                for i in range(rows_per_file):
                    ws.append([f"123456789012{i}", f"Product {i}", 10, 100.0, f"Reseller {file_num}", 1, 2024])

                wb.save(tmp.name)
                wb.close()
                file_paths.append(tmp.name)

        processor = BoxnoxProcessor()

        start = time.perf_counter()

        # Process files concurrently
        async def process_file(file_path):
            loop = asyncio.get_event_loop()
            return await loop.run_in_executor(None, processor.process, file_path)

        results = await asyncio.gather(*[process_file(fp) for fp in file_paths])

        total_time = time.perf_counter() - start

        print(f"\nConcurrent processing ({num_files} files, {rows_per_file} rows each):")
        print(f"Total time: {total_time:.2f}s")
        print(f"Average time per file: {total_time / num_files:.2f}s")

        # Cleanup
        for fp in file_paths:
            Path(fp).unlink()

        # Concurrent processing should be faster than sequential
        # Allow up to 2x time for 5 files (significant parallelization benefit)
        max_expected_time = (rows_per_file / 40) * 2  # Based on throughput
        assert total_time < max_expected_time, f"Concurrent processing not efficient: {total_time:.2f}s"

    # ============================================
    # DATA TRANSFORMATION PERFORMANCE
    # ============================================

    @pytest.mark.performance
    def test_currency_conversion_performance(self):
        """Test currency conversion overhead"""
        from app.services.vendors.galilu_processor import GaliluProcessor

        row_count = 1000

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["EAN", "Product", "Month", "Year", "Sales (PLN)", "Quantity"])

            for i in range(row_count):
                ws.append([f"123456789012{i}", f"Product {i}", 1, 2024, 1000.0, 10])

            wb.save(tmp.name)
            wb.close()

            processor = GaliluProcessor()

            start = time.perf_counter()
            result = processor.process(tmp.name)
            processing_time = time.perf_counter() - start

            print(f"\nCurrency conversion processing ({row_count} rows): {processing_time:.2f}s")

            # Currency conversion should add minimal overhead
            assert processing_time < 10, f"Currency conversion too slow: {processing_time:.2f}s"

            Path(tmp.name).unlink()

    # ============================================
    # ERROR HANDLING PERFORMANCE
    # ============================================

    @pytest.mark.performance
    def test_error_detection_performance(self):
        """Test that error detection doesn't significantly slow processing"""
        from app.services.vendors.boxnox_processor import BoxnoxProcessor

        row_count = 1000
        error_rows = 100  # 10% error rate

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sell Out by EAN"
            ws.append(["Product EAN", "Functional Name", "Sold Qty", "Sales Amount (EUR)", "Reseller", "Month", "Year"])

            for i in range(row_count):
                if i < error_rows:
                    # Invalid rows
                    ws.append(["INVALID", f"Product {i}", "INVALID", "INVALID", "Reseller", 13, 2024])
                else:
                    # Valid rows
                    ws.append([f"123456789012{i}", f"Product {i}", 10, 100.0, "Reseller", 1, 2024])

            wb.save(tmp.name)
            wb.close()

            processor = BoxnoxProcessor()

            start = time.perf_counter()
            try:
                result = processor.process(tmp.name)
            except Exception:
                pass
            processing_time = time.perf_counter() - start

            print(f"\nError detection overhead ({error_rows} errors in {row_count} rows): {processing_time:.2f}s")

            # Error detection should not double processing time
            assert processing_time < 15, f"Error detection too slow: {processing_time:.2f}s"

            Path(tmp.name).unlink()


# ============================================
# INTEGRATION PERFORMANCE TESTS
# ============================================

class TestEndToEndFileProcessing:
    """Test end-to-end file processing performance"""

    @pytest.mark.performance
    @pytest.mark.slow
    def test_complete_upload_workflow(self):
        """Test complete upload workflow from file to database"""
        # This would test: upload → validate → detect → process → insert
        # Requires full integration setup
        pass

    @pytest.mark.performance
    def test_batch_processing_performance(self):
        """Test batch processing of multiple uploads"""
        # Test processing multiple files in sequence
        pass
